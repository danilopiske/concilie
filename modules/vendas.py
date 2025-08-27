import streamlit as st
import pandas as pd
import os
from datetime import datetime
import time
from openpyxl import load_workbook
import numpy as np
import json
import sys
import traceback
from openpyxl import load_workbook

from proc.funcoesbd import (inserir_vendas_processadas, remover_vendas_duplicadas,buscar_ecs_por_cliente,
                            carregar_clientes, salvar_processamento, listar_processamentos_existentes,
                            gerar_novo_id_processamento, buscar_termos_filtraveis, buscar_bandeiras_por_ec,
                            inserir_vendas_filtradas)

ROOT_DIR = os.path.abspath(os.path.join(os.path.dirname(__file__), '..'))
sys.path.append(ROOT_DIR)

from proc.utils import (
    # carregar_clientes,
    # carregar_termos_filtraveis,
    # carregar_bandeiras_ativas,
    remover_acentos,
    limpar_forma_pgto,
    normalizar_bandeira,
    normalizar_forma_pagamento,
    limpar_forma_pgto
    # carregar_ecs_por_cliente
    # aplicar_conversoes
)

# Definições de caminho
PASTA_ENTRADA = "venda_planilhas"
# PASTA_SAIDA = "venda_planilhas"
# NOME_SAIDA = "vendas_consolidada.xlsx"
# NOME_FILTRADAS = "vendas_filtradas.xlsx"

os.makedirs(PASTA_ENTRADA, exist_ok=True)

def salvar_arquivo(uploaded_file):
    caminho_arquivo = os.path.join(PASTA_ENTRADA, uploaded_file.name)
    with open(caminho_arquivo, "wb") as f:
        f.write(uploaded_file.getbuffer())
    return caminho_arquivo

def encontrar_linha_cabecalho(df, colunas_esperadas, linha_min=2, linha_max=15, minimo_encontradas=4):
        for i in range(linha_min - 1, linha_max):
            linha = df.iloc[i].astype(str).str.strip().tolist()
            encontrados = sum(1 for col in colunas_esperadas if any(col.lower() == cel.lower() for cel in linha))
            if encontrados >= minimo_encontradas:
                return i
        return None

def carregar_excel_sanitizado(caminho_arquivo):
    wb = load_workbook(caminho_arquivo, data_only=True)
    ws = wb.active

    dados = []
    for row in ws.iter_rows(values_only=True):
        linha_sanitizada = []
        for cel in row:
            try:
                # Trata overflow ou converte tudo para string
                if isinstance(cel, float) and (cel == float('inf') or cel == float('-inf')):
                    linha_sanitizada.append("")
                else:
                    linha_sanitizada.append(str(cel) if cel is not None else "")
            except Exception:
                linha_sanitizada.append("")
        dados.append(linha_sanitizada)

    return pd.DataFrame(dados)
    
def processar_lancamentos(arquivos, cliente_id, ec_id, id_processamento):
    arquivos_processados = []

    if not arquivos:
        st.warning("⚠️ Nenhum arquivo enviado.")
        return None

    ecs = buscar_ecs_por_cliente(cliente_id)
    if not ecs:
        st.warning("⚠️ Nenhum EC encontrado para este cliente.")
        return None

    termos_filtraveis = []
    bandeiras_ativas = []

    for ec in ecs:
        termos_filtraveis += buscar_termos_filtraveis(ec)
        bandeiras_dict = buscar_bandeiras_por_ec(ec)
        bandeiras_ativas += [b for b, ativo in bandeiras_dict.items() if ativo == 1]

    termos_filtraveis = list(set(termos_filtraveis))
    bandeiras_ativas_upper = list(set(b.upper() for b in bandeiras_ativas))
        
    colunas_esperadas = [
       
        'Data da venda',
        'Hora da venda',
        'Estabelecimento',
        'CPF/CNPJ do estabelecimento',
        'Forma de pagamento',
        'Quantidade total de parcelas',
        'Bandeira'


    ]
    



    for arquivo in arquivos:
        nome_arquivo = arquivo.name
        inicio = time.time()
        

        try:
            caminho_arquivo = salvar_arquivo(arquivo)
            
            df_raw = carregar_excel_sanitizado(caminho_arquivo)
            linha_header = encontrar_linha_cabecalho(df_raw, colunas_esperadas)
            
            if linha_header is None:
                st.error(f"❌ Cabeçalho não identificado em {arquivo.name}.")
                continue

            df_raw = df_raw.loc[:, df_raw.iloc[linha_header].astype(str).str.strip() != ""]

            df_raw.columns = df_raw.iloc[linha_header]
            df = df_raw.iloc[linha_header + 1:].reset_index(drop=True)
            
            # Após normalizar, substitui
            # partes = df['Produto cielo'].str.partition(' ')
            # df['Bandeira'] = partes[0]
            # df['Forma de pagamento'] = partes[2]
            # df['Forma de pagamento'] = df['Forma de pagamento'].apply(normalizar_forma_pagamento)
            # df['Bandeira'] = df['Bandeira'].apply(normalizar_bandeira)
            # df['Forma de pagamento'] = df['Forma de pagamento'].apply(limpar_forma_pgto)
                      
            # Normalização
            colunas_str = df.select_dtypes(include='object').columns
            df[colunas_str] = df[colunas_str].applymap(lambda x: remover_acentos(x).upper() if isinstance(x, str) else x)

            # Filtros
            df['Filtrado por termo'] = df['Produto cielo'].str.lower().apply(
                lambda x: any(termo in x for termo in termos_filtraveis)
            )
            df['Filtrado por bandeira'] = ~df['Bandeira'].isin(bandeiras_ativas_upper)
            df['Filtrado'] = df['Filtrado por termo'] | df['Filtrado por bandeira']
            df['Tratar ou Ignorar'] = df['Filtrado'].map(lambda x: "IGNORAR" if x else "TRATAR")

            # Campos adicionais
            df['Data da autorização da venda'] = df['Data da Transação']
            df['Canal de venda'] = 'NÃO INFORMADO'
            df['Código da venda'] = 'NÃO INFORMADO'
            df['Número do cartão'] = 'NÃO INFORMADO'
            df['Tipo de captura'] = 'NÃO INFORMADO'
            df['Receba rápido'] = np.where(df['Percentual Des Prz Frexi'] != 0, 'Sim', 'Não')
            df['Taxas (%)'] = df['Percentual de Desconto']
            df['Taxas RR'] = df['Percentual Des Prz Frexi']
            df['Número da nota fiscal'] = '000000000'
            df['Taxa de embarque'] = '0,00'
            df['Comissão Mínima'] = '0,00'
            df['Status'] = 'APROVADA'
            df['Valor da entrada'] = '0,00'
            df['Valor do saque'] = '0,00'
            df['cliente_id'] = cliente_id
            df['ec_id'] = ec_id
            df['arquivo_origem'] = nome_arquivo
            df['id_processamento'] = id_processamento

            df = df.rename(columns={
                'Código Autorização (Transação)': 'Código de autorização',
                'Data da Transação': 'Data da venda',
                'Quantidade de Parcelas': 'Quantidade de parcelas',
                'Número RO': 'Resumo da operação',
                'Valor da Transação': 'Valor da venda',
                'Valor Comissão Bruta': 'Valor descontado',
                'Valor Líquido': 'Valor líquido da venda',
                'Número Referência Original (NSU)': 'NSU',
                'Data Crédito Ec': 'Previsão de pagamento',
                'Valor Comisssão Brt Prz Flexi': 'Valor RR',
                'Equipamento Lógico' : 'Número da máquina'
            })
        
            df['Quantidade de parcelas'] = df['Quantidade de parcelas'].replace(['', '0', 0, np.nan], 1).astype(int)
            df['Código de autorização'] = df['Código de autorização'].astype(str).str.zfill(6)
            df['Taxas (%)'] = pd.to_numeric(df['Taxas (%)'], errors='coerce').fillna(0)
            df['Taxas RR'] = pd.to_numeric(df['Taxas RR'], errors='coerce').fillna(0)
            df['Valor descontado'] = (
                df['Valor descontado'].astype(str).str.replace(',', '.', regex=False)
                .replace(['', 'nan', 'NaN', 'None'], '0')
                .str.replace(r'\s+', '', regex=True)
            )

            df['Valor descontado'] = pd.to_numeric(df['Valor descontado'], errors='coerce').fillna(0)
            df["Valor líquido da venda"] = pd.to_numeric(df["Valor líquido da venda"], errors='coerce').fillna(0).round(2)
            df['Previsão de pagamento'] = df['Previsão de pagamento'].replace('', np.nan)
            df['Previsão de pagamento'] = df['Previsão de pagamento'].fillna(df['Data da venda'])
            
            # Remove linhas com "TOTAL" ou "TOTAIS" na coluna 'Produto cielo' e campos vazios nas colunas essenciais
            colunas_essenciais = ['Bandeira', 'Forma de pagamento', 'Previsão de pagamento', 'Número da máquina']

            def linha_vazia(row):
                return all(
                    pd.isna(row[col]) or str(row[col]).strip() == ''
                    for col in colunas_essenciais
                )

            df = df[~df.apply(linha_vazia, axis=1)].reset_index(drop=True)


            nova_ordem = [
                'Data da venda', 'Data da autorização da venda', 'Bandeira', 'Forma de pagamento',
                'Quantidade de parcelas', 'Resumo da operação', 'Valor da venda', 'Taxas (%)', 'Valor descontado',
                'Taxas RR', 'Valor RR', 'Previsão de pagamento', 'Valor líquido da venda',
                'Canal de venda', 'Número da máquina', 'Código da venda', 'Código de autorização',
                'NSU', 'Número do cartão', 'Tipo de captura', 'Receba rápido', 'Comissão Mínima',
                'Número da nota fiscal', 'Taxa de embarque', 'Valor da entrada', 'Valor do saque',
                'Status', 'Tratar ou Ignorar', 'cliente_id', 'ec_id', 'arquivo_origem', 'id_processamento'
            ]
            
            df = df[nova_ordem + ['Filtrado']]
            df.reset_index(drop=True, inplace=True)
            df_tratadas = df[df['Filtrado'] == False].copy()
            df_filtradas = df[df['Filtrado'] == True].copy()

            qtde_tratadas = 0
            if not df_tratadas.empty:
                qtde_tratadas = inserir_vendas_processadas(df_tratadas)
            
            qtde_filtradas = 0
            if not df_filtradas.empty:
                qtde_filtradas = inserir_vendas_filtradas(df_filtradas)
            
            registros_total = qtde_tratadas + qtde_filtradas
            fim = time.time()
            duracao = fim - inicio
            arquivos_processados.append(
                f"✅ {registros_total} registros inseridos de `{nome_arquivo}` em {duracao:.2f} segundos."
            )
        except Exception as e:
                st.error(f"Erro ao processar {arquivo.name}: {e}")
    remover_vendas_duplicadas("vendas_processadas", id_processamento)
    remover_vendas_duplicadas("vendas_filtradas", id_processamento)
    
    return arquivos_processados

def interface():
    st.title("🧾 Processador de Lançamentos")

    # Carrega os clientes do banco
    lista_clientes = carregar_clientes()
    if not lista_clientes:
        st.warning("Nenhum cliente encontrado.")
        st.stop()

    # Mapeia para exibição no selectbox
    opcoes_clientes = {
        f"{c['cliente_id']} - {c['nome_fantasia']} - {c['cnpj']}": c['cliente_id']
        for c in lista_clientes
    }

    cliente_label = st.selectbox("👤 Selecione o cliente", list(opcoes_clientes))
    cliente_id = opcoes_clientes[cliente_label]

    # Busca ECs vinculados
    ecs = buscar_ecs_por_cliente(cliente_id)
    if not ecs:
        st.warning("Este cliente não possui ECs vinculados.")
        st.stop()

    ec_selecionado = st.selectbox("🏦 Escolha o EC", ecs)

    # Modo de processamento
    modo = st.radio("Modo de processamento", ["📦 Processamento novo", "🔄 Continuar existente"])

    descricao = ""
    id_processamento = ""

    if modo == "📦 Processamento novo":
        id_processamento, data_proc = gerar_novo_id_processamento(ec_selecionado)
        descricao = st.text_input(
            "Descrição do processamento",
            f"Processamento {data_proc.strftime('%d/%m/%Y %H:%M')}"
        )


    else:
        ids_existentes = listar_processamentos_existentes(ec_selecionado)
        if not ids_existentes:
            st.warning("Nenhum processamento existente encontrado para esse EC.")
            st.stop()
        id_processamento = st.selectbox("Selecione o processamento existente", ids_existentes)

    # Upload de arquivos
    uploaded_files = st.file_uploader(
        "📂 Envie arquivos Excel para processar", accept_multiple_files=True, type=["xlsx"]
    )

    # Armazena no session_state
    if uploaded_files:
        st.session_state.arquivos_vendas = uploaded_files

    arquivos = st.session_state.get("arquivos_vendas", [])

    # Botão de processamento
    if arquivos and st.button("📊 Processar Vendas"):
        resultado = processar_lancamentos(arquivos, cliente_id, ec_selecionado, id_processamento)
        st.write('Chegou antes aqui')
        if resultado:
            for msg in resultado:
                st.success(msg)
            # Salvar o processamento só após sucesso real
            if modo == "📦 Processamento novo":
                salvar_processamento(ec_selecionado, cliente_id, id_processamento, descricao, data_proc)
                st.write('Chegou aqui')
        else:
            st.error("❌ Nenhum arquivo foi processado.")
