"""Geração dos XLSX de conciliação Rede (4 arquivos separados, agrupados em um ZIP)."""

from __future__ import annotations

import io
import zipfile
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from app.services.conversor.rede_parser import ResultadoParsing

_AZUL_HEADER = "1e3a5f"
_FONT_HEADER = Font(bold=True, color="FFFFFF", name="Calibri", size=10)
_FILL_HEADER = PatternFill(start_color=_AZUL_HEADER, end_color=_AZUL_HEADER, fill_type="solid")
_ALIGN_CENTER = Alignment(horizontal="center", vertical="center")
_FMT_MOEDA = 'R$ #,##0.00'
_FMT_DATA = 'DD/MM/YYYY'


def _escrever_cabecalho(ws, colunas: list[str]) -> None:
    for col_idx, nome in enumerate(colunas, 1):
        cell = ws.cell(row=1, column=col_idx, value=nome)
        cell.font = _FONT_HEADER
        cell.fill = _FILL_HEADER
        cell.alignment = _ALIGN_CENTER
    ws.freeze_panes = "A2"


def _ajustar_larguras(ws, colunas: list[str]) -> None:
    for col_idx, nome in enumerate(colunas, 1):
        col_letter = get_column_letter(col_idx)
        max_len = max(len(nome), 12)
        for cell in ws.iter_cols(min_col=col_idx, max_col=col_idx, min_row=2):
            for c in cell:
                if c.value is not None:
                    val_len = len(str(c.value))
                    max_len = min(max(max_len, val_len), 40)
        ws.column_dimensions[col_letter].width = max_len + 2


def _formatar_colunas_valor(ws, indices_valor: list[int], indices_data: list[int]) -> None:
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            if cell.column in indices_valor:
                cell.number_format = _FMT_MOEDA
            elif cell.column in indices_data:
                cell.number_format = _FMT_DATA


def _wb_vendas_credito(resultados: list[ResultadoParsing]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas_Credito"
    cols = [
        "Estabelecimento", "Origem_Arquivo", "Data_Venda", "Data_Recebimento",
        "Resumo_Vendas", "Bandeira", "Quantidade", "Modalidade",
        "Valor_Bruto", "Valor_Correcao", "Valor_Liquido", "Tipo_Lancamento",
    ]
    _escrever_cabecalho(ws, cols)
    row = 2
    for res in resultados:
        for v in res.vendas_credito:
            ws.cell(row, 1, v.estabelecimento)
            ws.cell(row, 2, v.origem_arquivo)
            ws.cell(row, 3, v.data_venda)
            ws.cell(row, 4, v.data_recebimento)
            ws.cell(row, 5, v.resumo_vendas)
            ws.cell(row, 6, v.bandeira)
            ws.cell(row, 7, v.quantidade)
            ws.cell(row, 8, v.modalidade)
            ws.cell(row, 9, v.valor_bruto)
            ws.cell(row, 10, v.valor_correcao)
            ws.cell(row, 11, v.valor_liquido)
            ws.cell(row, 12, v.tipo_lancamento)
            row += 1
    _formatar_colunas_valor(ws, [9, 10, 11], [3, 4])
    _ajustar_larguras(ws, cols)
    return wb


def _wb_vendas_debito(resultados: list[ResultadoParsing]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas_Debito"
    cols = [
        "Estabelecimento", "Origem_Arquivo", "Data_Venda", "Data_Recebimento",
        "Resumo_Vendas", "Bandeira", "Quantidade", "Modalidade",
        "Valor_Bruto", "Valor_Saque", "Valor_Liquido", "Banco_Agencia_Conta",
    ]
    _escrever_cabecalho(ws, cols)
    row = 2
    for res in resultados:
        for v in res.vendas_debito:
            ws.cell(row, 1, v.estabelecimento)
            ws.cell(row, 2, v.origem_arquivo)
            ws.cell(row, 3, v.data_venda)
            ws.cell(row, 4, v.data_recebimento)
            ws.cell(row, 5, v.resumo_vendas)
            ws.cell(row, 6, v.bandeira)
            ws.cell(row, 7, v.quantidade)
            ws.cell(row, 8, v.modalidade)
            ws.cell(row, 9, v.valor_bruto)
            ws.cell(row, 10, v.valor_saque)
            ws.cell(row, 11, v.valor_liquido)
            ws.cell(row, 12, v.banco_agencia_conta)
            row += 1
    _formatar_colunas_valor(ws, [9, 10, 11], [3, 4])
    _ajustar_larguras(ws, cols)
    return wb


def _wb_pagamentos(resultados: list[ResultadoParsing]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamentos"
    cols = [
        "Estabelecimento", "Origem_Arquivo", "Data_Recebimento",
        "Ordem_Credito", "Valor_Liquido", "Banco_Agencia_Conta",
    ]
    _escrever_cabecalho(ws, cols)
    row = 2
    for res in resultados:
        for p in res.pagamentos:
            ws.cell(row, 1, p.estabelecimento)
            ws.cell(row, 2, p.origem_arquivo)
            ws.cell(row, 3, p.data_recebimento)
            ws.cell(row, 4, p.ordem_credito)
            ws.cell(row, 5, p.valor_liquido)
            ws.cell(row, 6, p.banco_agencia_conta)
            row += 1
    _formatar_colunas_valor(ws, [5], [3])
    _ajustar_larguras(ws, cols)
    return wb


def _wb_recebiveis(resultados: list[ResultadoParsing]) -> Workbook:
    """Antiga aba 'Tarifas_e_Debitos', renomeada para 'Recebiveis'."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Recebiveis"
    cols = [
        "Estabelecimento", "Origem_Arquivo", "Data_Inclusao", "Data_Pagamento",
        "Motivo_Debito", "Resumo", "Valor_Devido", "Valor_Debitado", "Meio_Pagamento",
    ]
    _escrever_cabecalho(ws, cols)
    row = 2
    for res in resultados:
        for t in res.tarifas_debitos:
            ws.cell(row, 1, t.estabelecimento)
            ws.cell(row, 2, t.origem_arquivo)
            ws.cell(row, 3, t.data_inclusao)
            ws.cell(row, 4, t.data_pagamento)
            ws.cell(row, 5, t.motivo_debito)
            ws.cell(row, 6, t.resumo)
            ws.cell(row, 7, t.valor_devido)
            ws.cell(row, 8, t.valor_debitado)
            ws.cell(row, 9, t.meio_pagamento)
            row += 1
    _formatar_colunas_valor(ws, [7, 8], [3, 4])
    _ajustar_larguras(ws, cols)
    return wb


def _wb_para_bytes(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.read()


def gerar_zip(resultados: list[ResultadoParsing]) -> bytes:
    """Gera um ZIP contendo os 4 XLSX separados (Vendas_Credito, Vendas_Debito,
    Pagamentos, Recebiveis) e retorna como bytes."""
    arquivos = {
        "Vendas_Credito.xlsx": _wb_para_bytes(_wb_vendas_credito(resultados)),
        "Vendas_Debito.xlsx": _wb_para_bytes(_wb_vendas_debito(resultados)),
        "Pagamentos.xlsx": _wb_para_bytes(_wb_pagamentos(resultados)),
        "Recebiveis.xlsx": _wb_para_bytes(_wb_recebiveis(resultados)),
    }

    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as zf:
        for nome, conteudo in arquivos.items():
            zf.writestr(nome, conteudo)
    buf.seek(0)
    return buf.read()


def nome_arquivo_saida(resultados: list[ResultadoParsing]) -> str:
    if len(resultados) == 1 and resultados[0].periodo:
        periodo = resultados[0].periodo.replace(' ', '_').replace('/', '-')
        return f"Conciliacao_Rede_{periodo}.zip"
    return "Conciliacao_Rede_Consolidado.zip"
