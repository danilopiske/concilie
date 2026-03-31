"""
Serviço de pré-processamento de relatórios.

Fluxo:
1. preprocessar_relatorio() → calcula cada seção → salva .parquet + _meta.json
2. invalidar_parquet()      → deleta pasta parquet_cache/{processamento_id}/
3. emitir_modelo()          → lê parquets necessários → renderiza template

REGRA CRÍTICA: gerar_relatorio_html() em modules/reports.py NÃO é modificado.
Este serviço reutiliza as funções de cálculo existentes via importação.
"""

import json
import logging
import os
import shutil
from datetime import datetime
from typing import Optional

import pandas as pd
import polars as pl
from jinja2 import Environment, FileSystemLoader
from sqlalchemy.engine import Engine

from app.core.database import SessionLocal
from app.models.modelo_relatorio import ModeloRelatorio

logger = logging.getLogger(__name__)

# Diretório base dos parquets de pré-processamento
_PREPROC_CACHE_DIR = os.path.join(
    os.path.dirname(__file__), "..", "..", "..", "..", "apps", "api", "parquet_cache"
)
_PREPROC_CACHE_DIR = os.path.normpath(_PREPROC_CACHE_DIR)

# Diretório dos templates HTML/XML
_TEMPLATES_DIR = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "..", "..", "..", "templates")
)


# ---------------------------------------------------------------------------
# Helpers de path
# ---------------------------------------------------------------------------

def _safe(s: str) -> str:
    return "".join(c if c.isalnum() or c in "_-" else "_" for c in s)


def _adq_slug(adquirente: Optional[str]) -> str:
    """Normaliza o adquirente para nome de subdiretório. None/"Todos"/"" → 'todos'."""
    if not adquirente or adquirente.strip().lower() in ("todos", "all", ""):
        return "todos"
    return _safe(adquirente.strip().lower())


def _pasta_processamento(processamento_id: str, adquirente: Optional[str] = None) -> str:
    base = os.path.join(_PREPROC_CACHE_DIR, _safe(processamento_id))
    return os.path.join(base, _adq_slug(adquirente))


def _parquet_path(processamento_id: str, secao: str, adquirente: Optional[str] = None) -> str:
    return os.path.join(_pasta_processamento(processamento_id, adquirente), f"{secao}.parquet")


def _meta_path(processamento_id: str, adquirente: Optional[str] = None) -> str:
    return os.path.join(_pasta_processamento(processamento_id, adquirente), "_meta.json")


# ---------------------------------------------------------------------------
# Status do cache
# ---------------------------------------------------------------------------

def status_parquet(processamento_id: str, adquirente: Optional[str] = None) -> dict:
    """
    Retorna se o parquet existe para o combo (processamento_id + adquirente).
    adquirente=None/"Todos" → verifica o slot 'todos'.
    Fallback: se slug não existe, verifica o cache root antigo (sem slug).
    """
    meta = _meta_path(processamento_id, adquirente)
    if not os.path.exists(meta):
        # Fallback para cache root antigo (formato anterior à migração por adquirente)
        root_meta = os.path.join(_PREPROC_CACHE_DIR, _safe(processamento_id), "_meta.json")
        if os.path.exists(root_meta):
            meta = root_meta
    if os.path.exists(meta):
        with open(meta, "r", encoding="utf-8") as f:
            dados = json.load(f)
        return {
            "existe": True,
            "gerado_em": dados.get("gerado_em"),
            "calc_tipo": dados.get("calc_tipo"),
            "adquirente": dados.get("adquirente"),
        }
    return {"existe": False, "gerado_em": None, "calc_tipo": None, "adquirente": None}


# ---------------------------------------------------------------------------
# Invalidação
# ---------------------------------------------------------------------------

def invalidar_parquet(processamento_id: str, adquirente: Optional[str] = None) -> None:
    """
    Deleta a pasta de parquets do combo (processamento_id + adquirente).
    Se adquirente=None, deleta TODA a pasta base do processamento (todos os slots).
    """
    if adquirente is None:
        # Apaga a pasta base inteira (todos os adquirentes)
        base = os.path.join(_PREPROC_CACHE_DIR, _safe(processamento_id))
        pasta = base
    else:
        pasta = _pasta_processamento(processamento_id, adquirente)

    if os.path.exists(pasta):
        shutil.rmtree(pasta)
        logger.info(f"[PREPROC] Cache parquet invalidado: {pasta}")
    else:
        logger.debug(f"[PREPROC] Nenhum cache para invalidar: {pasta}")


# ---------------------------------------------------------------------------
# Pré-processamento
# ---------------------------------------------------------------------------

def preprocessar_relatorio(
    engine: Engine,
    processamento_id: str,
    calc_tipo: Optional[str] = None,
    data_inicio: Optional[datetime] = None,
    data_fim: Optional[datetime] = None,
    adquirente: Optional[str] = None,
) -> dict:
    """
    Executa o pré-processamento completo: calcula todas as seções e salva
    como .parquet individuais em parquet_cache/{processamento_id}/{adquirente_slug}/.

    adquirente=None/"Todos" → slot 'todos' (sem filtro por adquirente).
    adquirente="Cielo"     → slot 'cielo' (filtrado por Cielo).

    Reutiliza as funções de cálculo existentes de modules/reports.py
    sem modificá-las.

    Retorna dict com status e seções geradas.
    """
    # Importação local para evitar circular import e manter isolamento
    from modules.reports import (
        load_vendas_calculos_cached,
        calcular_perdas_por_semestre,
        calcular_min_max_taxas_agrupado,
        calcular_contagem_taxas_agrupado,
        calcular_sumario_recebiveis,
        obter_dados_bancarios_distintos,
        obter_evidencias_transacoes,
        filtrar_valores_rede_depara,
        calcular_previsao_pagamento_rede,
        _get_base_id,
    )

    pasta = _pasta_processamento(processamento_id, adquirente)
    os.makedirs(pasta, exist_ok=True)

    # Normaliza adquirente para filtro
    adq_filtro = None if not adquirente or adquirente.strip().lower() in ("todos", "all", "") else adquirente.strip()

    secoes_geradas = []
    erros = []

    # 1. Carregar dataset principal (Polars) — mesmo fluxo do gerar_relatorio_html
    logger.info(f"[PREPROC] Carregando dataset principal para {processamento_id}")
    df_cached = load_vendas_calculos_cached(engine, processamento_id, calc_tipo)

    if df_cached.is_empty() and calc_tipo:
        base_id = _get_base_id(processamento_id)
        from modules.reports import read_sql_polars, _convert_placeholders
        sql_fallback = (
            "SELECT id_venda, data_venda, bandeira, forma_pagamento, tx_rr_venda, "
            "vl_rr_venda, vl_venda, tx_venda, desc_venda, vl_liq_venda, tx_calc, "
            "desc_calc, vl_liq_calc, perda, adquirente, nsu, cod_autorizacao, "
            "perda_rr, ec_id FROM vendas_calculos WHERE calc_id LIKE %s"
        )
        df_cached = read_sql_polars(sql_fallback, engine, params=(f"{base_id}%",))

    # Aplicar filtros de data e adquirente
    if not df_cached.is_empty():
        lf = df_cached.lazy()
        lf = filtrar_valores_rede_depara(lf)
        lf = calcular_previsao_pagamento_rede(lf)
        cols = lf.collect_schema().names()
        data_col = next((c for c in ["Data_da_venda", "data_venda", "Data"] if c in cols), "data_venda")
        if data_inicio:
            lf = lf.filter(pl.col(data_col).cast(pl.Date) >= pl.lit(data_inicio).cast(pl.Date))
        if data_fim:
            lf = lf.filter(pl.col(data_col).cast(pl.Date) <= pl.lit(data_fim).cast(pl.Date))
        if adq_filtro:
            adq_col = next((c for c in cols if "adquirente" in c.lower()), None)
            if adq_col:
                lf = lf.filter(pl.col(adq_col).cast(pl.Utf8).str.to_lowercase() == adq_filtro.lower())
        df_main = lf.collect()
    else:
        df_main = df_cached

    # 2. Seção: vendas_calculos
    # Calcular stats reais sobre df_main completo ANTES de truncar
    _total_transacoes_real = len(df_main) if not df_main.is_empty() else 0
    _stats_reais: dict = {}
    if not df_main.is_empty():
        try:
            _cols = df_main.collect_schema().names()
            _col_vl   = next((c for c in _cols if c.lower() in ("vl_venda", "valor_venda")), None)
            _col_liq  = next((c for c in _cols if c.lower() in ("vl_liq_calc", "vl_liq_venda")), None)
            _col_tx   = next((c for c in _cols if c.lower() in ("tx_calc", "tx_venda")), None)
            _col_vmed = next((c for c in _cols if c.lower() in ("vl_venda", "valor_venda")), None)
            _df_agg = df_main.select([
                pl.col(_col_vl).sum().alias("fat")  if _col_vl  else pl.lit(0.0).alias("fat"),
                pl.col(_col_liq).sum().alias("liq") if _col_liq else pl.lit(0.0).alias("liq"),
                pl.col(_col_tx).mean().alias("tm")  if _col_tx  else pl.lit(0.0).alias("tm"),
                pl.col(_col_tx).min().alias("tmin") if _col_tx  else pl.lit(0.0).alias("tmin"),
                pl.col(_col_tx).max().alias("tmax") if _col_tx  else pl.lit(0.0).alias("tmax"),
                pl.col(_col_vl).mean().alias("vmedio") if _col_vl else pl.lit(0.0).alias("vmedio"),
                pl.col(_col_vl).min().alias("vmin")   if _col_vl else pl.lit(0.0).alias("vmin"),
                pl.col(_col_vl).max().alias("vmax")   if _col_vl else pl.lit(0.0).alias("vmax"),
            ]).row(0, named=True)
            _stats_reais = {
                "_total_transacoes_real": _total_transacoes_real,
                "_faturamento_bruto_real": float(_df_agg["fat"] or 0),
                "_valor_liquido_real":     float(_df_agg["liq"] or 0),
                "_taxa_media_real":        float(_df_agg["tm"]  or 0),
                "_taxa_min_real":          float(_df_agg["tmin"] or 0),
                "_taxa_max_real":          float(_df_agg["tmax"] or 0),
                "_valor_medio_real":       float(_df_agg["vmedio"] or 0),
                "_valor_min_real":         float(_df_agg["vmin"] or 0),
                "_valor_max_real":         float(_df_agg["vmax"] or 0),
            }
        except Exception as _e:
            logger.warning(f"[PREPROC] stats reais erro: {_e}")
            _stats_reais = {"_total_transacoes_real": _total_transacoes_real}

    try:
        if not df_main.is_empty():
            _limit = 100_000
            df_save = df_main.head(_limit) if len(df_main) > _limit else df_main
            df_save_pd = df_save.to_pandas()
            for _k, _v in _stats_reais.items():
                df_save_pd[_k] = _v
            pl.from_pandas(df_save_pd).write_parquet(_parquet_path(processamento_id, "vendas_calculos", adquirente))
            secoes_geradas.append("vendas_calculos")
            logger.info(f"[PREPROC] vendas_calculos: {len(df_save)} linhas (total real: {_total_transacoes_real})")
    except Exception as e:
        erros.append(f"vendas_calculos: {e}")
        logger.error(f"[PREPROC] Erro em vendas_calculos: {e}")

    # 3. Seção: perdas_semestre
    try:
        df_perdas = calcular_perdas_por_semestre(df_main, incluir_faturamento=True)
        if not df_perdas.empty:
            pl.from_pandas(df_perdas).write_parquet(_parquet_path(processamento_id, "perdas_semestre", adquirente))
            secoes_geradas.append("perdas_semestre")
    except Exception as e:
        erros.append(f"perdas_semestre: {e}")
        logger.error(f"[PREPROC] Erro em perdas_semestre: {e}")

    # 4. Seção: taxas_minmax
    try:
        df_taxas = calcular_min_max_taxas_agrupado(df_main)
        if not df_taxas.empty:
            pl.from_pandas(df_taxas).write_parquet(_parquet_path(processamento_id, "taxas_minmax", adquirente))
            secoes_geradas.append("taxas_minmax")
    except Exception as e:
        erros.append(f"taxas_minmax: {e}")
        logger.error(f"[PREPROC] Erro em taxas_minmax: {e}")

    # 5. Seção: contagem_transacoes
    try:
        df_contagem = calcular_contagem_taxas_agrupado(df_main)
        if not df_contagem.empty:
            pl.from_pandas(df_contagem).write_parquet(_parquet_path(processamento_id, "contagem_transacoes", adquirente))
            secoes_geradas.append("contagem_transacoes")
    except Exception as e:
        erros.append(f"contagem_transacoes: {e}")
        logger.error(f"[PREPROC] Erro em contagem_transacoes: {e}")

    # 6. Seção: recebiveis_sumario
    try:
        df_rec = calcular_sumario_recebiveis(engine, processamento_id, data_inicio, data_fim)
        if not df_rec.empty:
            pl.from_pandas(df_rec).write_parquet(_parquet_path(processamento_id, "recebiveis_sumario", adquirente))
            secoes_geradas.append("recebiveis_sumario")
    except Exception as e:
        erros.append(f"recebiveis_sumario: {e}")
        logger.error(f"[PREPROC] Erro em recebiveis_sumario: {e}")

    # 7. Seção: dados_bancarios
    try:
        df_banco = obter_dados_bancarios_distintos(engine, processamento_id, data_inicio, data_fim)
        if df_banco is not None and not df_banco.empty:
            pl.from_pandas(df_banco).write_parquet(_parquet_path(processamento_id, "dados_bancarios", adquirente))
            secoes_geradas.append("dados_bancarios")
    except Exception as e:
        erros.append(f"dados_bancarios: {e}")
        logger.error(f"[PREPROC] Erro em dados_bancarios: {e}")

    # 8. Seção: vendas_filtradas
    try:
        from modules.reports import read_sql_polars, _get_base_id as _gbid2
        _base_id2 = _gbid2(processamento_id)
        _sql_vf = (
            "SELECT * FROM vendas_filtradas WHERE processamentoid LIKE %s"
        )
        df_vf = read_sql_polars(_sql_vf, engine, params=(f"{_base_id2}%",)).to_pandas()
        if len(df_vf) > 100_000:
            logger.warning(f"[PREPROC] Truncando vendas_filtradas para 100k (tinha {len(df_vf)})")
            df_vf = df_vf.head(100_000)
            
        if adq_filtro and not df_vf.empty:
            _col_adq = next((c for c in df_vf.columns if "adquirente" in c.lower()), None)
            if _col_adq:
                df_vf = df_vf[df_vf[_col_adq].astype(str).str.lower() == adq_filtro.lower()]
        if not df_vf.empty:
            pl.from_pandas(df_vf).write_parquet(_parquet_path(processamento_id, "vendas_filtradas", adquirente))
            secoes_geradas.append("vendas_filtradas")
    except Exception as e:
        erros.append(f"vendas_filtradas: {e}")
        logger.error(f"[PREPROC] Erro em vendas_filtradas: {e}")

    # 9. Seção: recebiveis_filtrados
    try:
        _sql_rf = (
            "SELECT * FROM recebiveis_filtrados WHERE processamentoid LIKE %s"
        )
        df_rfp = read_sql_polars(_sql_rf, engine, params=(f"{_base_id2}%",)).to_pandas()
        if len(df_rfp) > 100_000:
            logger.warning(f"[PREPROC] Truncando recebiveis_filtrados para 100k (tinha {len(df_rfp)})")
            df_rfp = df_rfp.head(100_000)
            
        if not df_rfp.empty:
            pl.from_pandas(df_rfp).write_parquet(_parquet_path(processamento_id, "recebiveis_filtrados", adquirente))
            secoes_geradas.append("recebiveis_filtrados")
    except Exception as e:
        erros.append(f"recebiveis_filtrados: {e}")
        logger.error(f"[PREPROC] Erro em recebiveis_filtrados: {e}")

    # 10. Seção: evidencias
    try:
        evidencias = obter_evidencias_transacoes(engine, processamento_id, calc_tipo, df=df_main)
        ev_frames = {}
        for chave, df_ev in evidencias.items():
            if df_ev is not None and not df_ev.empty:
                ev_frames[chave] = df_ev.to_dict(orient="records")
        if ev_frames:
            ev_path = _parquet_path(processamento_id, "evidencias", adquirente).replace(".parquet", ".json")
            with open(ev_path, "w", encoding="utf-8") as f:
                json.dump(ev_frames, f, ensure_ascii=False, default=str)
            secoes_geradas.append("evidencias")
    except Exception as e:
        erros.append(f"evidencias: {e}")
        logger.error(f"[PREPROC] Erro em evidencias: {e}")

    # Salvar _meta.json
    meta = {
        "gerado_em": datetime.now().isoformat(),
        "calc_tipo": calc_tipo or "",
        "adquirente": adq_filtro or "todos",
        "secoes_geradas": secoes_geradas,
    }
    with open(_meta_path(processamento_id, adquirente), "w", encoding="utf-8") as f:
        json.dump(meta, f, ensure_ascii=False)

    logger.info(f"[PREPROC] Concluído: {len(secoes_geradas)} seções. Erros: {len(erros)}")
    return {"secoes_geradas": secoes_geradas, "erros": erros, "meta": meta}


# ---------------------------------------------------------------------------
# Emissão de modelo
# ---------------------------------------------------------------------------

def _carregar_parquet(processamento_id: str, secao: str, adquirente: Optional[str] = None) -> pd.DataFrame:
    path = _parquet_path(processamento_id, secao, adquirente)
    if os.path.exists(path):
        return pl.read_parquet(path).to_pandas()
    # Fallback: cache antigo sem slug (antes da migração por adquirente)
    root_path = os.path.join(_PREPROC_CACHE_DIR, _safe(processamento_id), f"{secao}.parquet")
    if os.path.exists(root_path):
        return pl.read_parquet(root_path).to_pandas()
    return pd.DataFrame()


def _carregar_evidencias(processamento_id: str, adquirente: Optional[str] = None) -> dict:
    path = _parquet_path(processamento_id, "evidencias", adquirente).replace(".parquet", ".json")
    if os.path.exists(path):
        with open(path, "r", encoding="utf-8") as f:
            return json.load(f)
    return {}


def emitir_modelo(
    processamento_id: str,
    modelo_id: int,
    engine: Engine,
    output_dir: Optional[str] = None,
    opcoes: Optional[dict] = None,
) -> str:
    """
    Lê os parquets necessários para o modelo e renderiza o template.
    Retorna o caminho do arquivo gerado.
    """
    db = SessionLocal()
    try:
        modelo = db.query(ModeloRelatorio).filter(
            ModeloRelatorio.id == modelo_id,
            ModeloRelatorio.ativo == True,
        ).first()
        if not modelo:
            raise ValueError(f"Modelo {modelo_id} não encontrado ou inativo")

        secoes = json.loads(modelo.secoes_necessarias)
        template_arquivo = modelo.template_arquivo
        tipo = modelo.tipo
    finally:
        db.close()

    opcoes = opcoes or {}

    # Adquirente define qual slot de parquet carregar
    adquirente = opcoes.get("adquirente") or None

    # Carregar DataFrames das seções necessárias
    dados = {}
    for secao in secoes:
        if secao == "evidencias":
            dados["evidencias"] = _carregar_evidencias(processamento_id, adquirente)
        else:
            dados[secao] = _carregar_parquet(processamento_id, secao, adquirente)

    # Renderizar template
    env = Environment(loader=FileSystemLoader(_TEMPLATES_DIR))
    env.tests["number"] = lambda v: isinstance(v, (int, float))

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    safe_id = "".join(c if c.isalnum() or c in "_-" else "_" for c in processamento_id)
    nome_modelo_safe = "".join(c if c.isalnum() else "_" for c in modelo.nome.lower())

    if output_dir is None:
        from modules.reports import criar_diretorio_relatorios
        output_dir = criar_diretorio_relatorios()

    if tipo == "xml":
        # Excel via openpyxl (.xlsx)
        output_path = os.path.join(output_dir, f"relatorio_{safe_id}_{nome_modelo_safe}_{timestamp}.xlsx")
        _gerar_excel_xlsx(dados, processamento_id, engine, output_path, opcoes)
    elif "sintetico" in (template_arquivo or ""):
        # Sintético usa contexto próprio com KPIs calculados
        contexto = _montar_contexto_sintetico(dados, processamento_id, engine, opcoes)
        template = env.get_template(template_arquivo)
        html_str = template.render(contexto)
        ext = "html"
        output_path = os.path.join(output_dir, f"relatorio_{safe_id}_{nome_modelo_safe}_{timestamp}.{ext}")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_str)
    else:
        # HTML analítico/sem_capa via Jinja2
        contexto = _montar_contexto_html(dados, processamento_id, engine, opcoes)
        template = env.get_template(template_arquivo)
        html_str = template.render(contexto)
        ext = "html"
        output_path = os.path.join(output_dir, f"relatorio_{safe_id}_{nome_modelo_safe}_{timestamp}.{ext}")
        with open(output_path, "w", encoding="utf-8") as f:
            f.write(html_str)

    logger.info(f"[PREPROC] Modelo '{modelo.nome}' emitido: {output_path}")
    return output_path


def _montar_contexto_sintetico(dados: dict, processamento_id: str, engine: Engine, opcoes: Optional[dict] = None) -> dict:
    """
    Monta contexto para o template sintético, calculando KPIs executivos
    a partir dos parquets já carregados.
    """
    from modules.reports import obter_dados_processamento

    opcoes = opcoes or {}
    apenas_com_perdas = opcoes.get("apenas_com_perdas", False)
    adquirente_filtro = opcoes.get("adquirente") or None
    if adquirente_filtro and adquirente_filtro.lower() in ("todos", "all", ""):
        adquirente_filtro = None

    # Aplicar filtro de adquirente em vendas_calculos
    if adquirente_filtro:
        df_vc = dados.get("vendas_calculos")
        if df_vc is not None and not df_vc.empty:
            col_adq = next((c for c in df_vc.columns if "adquirente" in c.lower()), None)
            if col_adq:
                dados = dict(dados)
                dados["vendas_calculos"] = df_vc[df_vc[col_adq].astype(str).str.lower() == adquirente_filtro.lower()]

    ctx: dict = {}

    # Dados do processamento (cliente_nome, periodo, adquirente, ...)
    try:
        ctx.update(obter_dados_processamento(engine, processamento_id))
    except Exception:
        pass
    if adquirente_filtro:
        ctx["adquirente"] = adquirente_filtro

    ctx.setdefault("cliente_nome", "—")
    ctx.setdefault("periodo", "—")
    ctx.setdefault("adquirente", "Todos")
    ctx["data_geracao"] = datetime.now().strftime("%d/%m/%Y %H:%M")
    ctx["disclaimer"] = "Todas as análises são baseadas exclusivamente nos dados fornecidos pela Adquirente."

    def _fmt_brl(v: float) -> str:
        return f"R$ {v:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")

    def _fmt_pct(v: float) -> str:
        return f"{v:.2f}".replace(".", ",")

    # KPIs a partir de vendas_calculos
    df_vc = dados.get("vendas_calculos")
    total_transacoes = 0
    faturamento_bruto = 0.0
    valor_liquido = 0.0
    taxa_media = 0.0
    bandeiras: list = []
    top_valores: list = []

    if df_vc is not None and not df_vc.empty:
        def _real(col, fallback):
            return df_vc[col].iloc[0] if col in df_vc.columns else fallback
        total_transacoes = int(_real("_total_transacoes_real", len(df_vc)))

        # Detectar colunas de valor bruto: vl_venda (nome real no banco)
        col_bruto = next((c for c in df_vc.columns if c.lower() in ("vl_venda", "valor_venda", "vl_bruto")), None)
        if not col_bruto:
            col_bruto = next((c for c in df_vc.columns if "bruto" in c.lower() and "valor" in c.lower()), None)
        if not col_bruto:
            col_bruto = next((c for c in df_vc.columns if c.lower().startswith("vl_") and "venda" in c.lower()), None)

        # Detectar coluna de valor líquido: vl_liq_calc ou vl_liq_venda
        col_liquido = next((c for c in df_vc.columns if c.lower() in ("vl_liq_calc", "vl_liq_venda")), None)
        if not col_liquido:
            col_liquido = next((c for c in df_vc.columns if "liquid" in c.lower() and "valor" in c.lower()), None)

        # Detectar coluna de taxa: tx_calc ou tx_venda
        col_taxa = next((c for c in df_vc.columns if c.lower() in ("tx_calc", "tx_venda")), None)
        if not col_taxa:
            col_taxa = next((c for c in df_vc.columns if "taxa" in c.lower() and c.lower() not in ("taxa_contratada",)), None)

        col_bandeira = next((c for c in df_vc.columns if "bandeira" in c.lower()), None)
        col_data = next((c for c in df_vc.columns if "data" in c.lower()), None)
        col_nsu = next((c for c in df_vc.columns if c.lower() == "nsu"), None)
        col_cod_aut = next((c for c in df_vc.columns if c.lower() in ("cod_autorizacao", "cod_autorização", "autorizacao")), None)

        # Usar stats reais (calculadas antes do cap de 100k) se disponíveis
        faturamento_bruto = float(_real("_faturamento_bruto_real", df_vc[col_bruto].sum() if col_bruto else 0.0))
        valor_liquido     = float(_real("_valor_liquido_real",     df_vc[col_liquido].sum() if col_liquido else 0.0))
        taxa_media        = float(_real("_taxa_media_real",        df_vc[col_taxa].mean() if col_taxa else 0.0))

        # Bandeiras
        if col_bandeira and col_bruto:
            grp = df_vc.groupby(col_bandeira).agg(
                qtd=(col_bruto, "count"),
                valor=(col_bruto, "sum"),
            ).reset_index()
            total_b = grp["valor"].sum() or 1
            for _, row in grp.iterrows():
                bandeiras.append({
                    "nome": str(row[col_bandeira]),
                    "qtd": int(row["qtd"]),
                    "valor": _fmt_brl(float(row["valor"])),
                    "percentual": _fmt_pct(float(row["valor"]) / total_b * 100),
                })
            bandeiras.sort(key=lambda x: x["qtd"], reverse=True)

        # Top 3 maiores valores
        if col_bruto and col_data and col_bandeira and col_taxa:
            cols_top = [c for c in [col_data, col_bandeira, col_nsu, col_cod_aut, col_bruto, col_taxa] if c]
            top_df = df_vc.nlargest(3, col_bruto)[cols_top]
            for _, row in top_df.iterrows():
                top_valores.append({
                    "data": str(row[col_data])[:10],
                    "bandeira": str(row[col_bandeira]),
                    "nsu": str(row[col_nsu]) if col_nsu else "",
                    "cod_autorizacao": str(row[col_cod_aut]) if col_cod_aut else "",
                    "valor": _fmt_brl(float(row[col_bruto])),
                    "taxa": _fmt_pct(float(row[col_taxa])),
                })

    ticket_medio = (faturamento_bruto / total_transacoes) if total_transacoes else 0.0
    percentual_liquido = (valor_liquido / faturamento_bruto * 100) if faturamento_bruto else 0.0

    ctx["total_transacoes"] = f"{total_transacoes:,}".replace(",", ".")
    ctx["faturamento_bruto"] = _fmt_brl(faturamento_bruto)
    ctx["valor_liquido"] = _fmt_brl(valor_liquido)
    ctx["percentual_liquido"] = _fmt_pct(percentual_liquido)
    ctx["ticket_medio"] = _fmt_brl(ticket_medio)
    ctx["taxa_media"] = _fmt_pct(taxa_media)
    ctx["bandeiras"] = bandeiras
    ctx["top_valores"] = top_valores

    # Divergências: Perda MDR + Perda RR/RA (perdas_semestre) + recebíveis contestáveis
    def _to_float_sint(v) -> float:
        try:
            if isinstance(v, (int, float)):
                return float(v)
            s = str(v).replace("R$", "").replace(".", "").replace(",", ".").strip()
            return float(s)
        except Exception:
            return 0.0

    total_divergencias_num = 0.0
    df_perdas = dados.get("perdas_semestre")
    if df_perdas is not None and not df_perdas.empty:
        col_sem = next((c for c in df_perdas.columns if "semestre" in c.lower() or c == "Ano-Semestre"), None)
        df_rows = df_perdas[~df_perdas[col_sem].astype(str).str.contains("TOTAL", case=False, na=False)] if col_sem else df_perdas
        col_mdr = next((c for c in df_rows.columns if "mdr" in c.lower()), None)
        col_rr  = next((c for c in df_rows.columns if ("rr" in c.lower() or "ra" in c.lower()) and "perda" in c.lower()), None)
        if col_mdr:
            total_divergencias_num += abs(sum(_to_float_sint(v) for v in df_rows[col_mdr]))
        if col_rr:
            total_divergencias_num += abs(sum(_to_float_sint(v) for v in df_rows[col_rr]))
        # Fallback: se não encontrou MDR/RR separados, usa Perda Total
        if total_divergencias_num == 0.0:
            col_pt = next((c for c in df_rows.columns if c.lower() == "perda total"), None)
            if col_pt:
                total_divergencias_num = abs(sum(_to_float_sint(v) for v in df_rows[col_pt]))

    # Adicionar total de recebíveis com descontos contestáveis
    df_rec_sint = dados.get("recebiveis_sumario")
    if df_rec_sint is not None and not df_rec_sint.empty:
        col_sem_rs = next((c for c in df_rec_sint.columns if "semestre" in c.lower() or c == "Ano-Semestre"), None)
        df_rec_rows_s = df_rec_sint[~df_rec_sint[col_sem_rs].astype(str).str.contains("TOTAL", case=False, na=False)] if col_sem_rs else df_rec_sint
        col_val_rs = next((c for c in df_rec_rows_s.columns if "valor" in c.lower()), None)
        if col_val_rs:
            rec_vals_s = pd.to_numeric(df_rec_rows_s[col_val_rs], errors="coerce").fillna(0)
            total_divergencias_num += float(rec_vals_s.sum())

    ctx["total_divergencias_num"] = total_divergencias_num
    ctx["total_divergencias"] = _fmt_brl(total_divergencias_num)

    # Resumo faturamento (texto) com período e contagem de dias
    periodo_str = ctx.get("periodo", "")
    try:
        from modules.reports import obter_adquirentes_e_periodo_processamento
        _p = obter_adquirentes_e_periodo_processamento(engine, processamento_id)
        _dmin = _p.get("data_min")
        _dmax = _p.get("data_max")
        if _dmin and _dmax:
            import pandas as _pd_tmp
            _d0 = _pd_tmp.to_datetime(_dmin)
            _d1 = _pd_tmp.to_datetime(_dmax)
            _dias = (_d1 - _d0).days + 1
            periodo_str = f"{_d0.strftime('%d/%m/%Y')} a {_d1.strftime('%d/%m/%Y')} ({_dias} dias)"
    except Exception:
        pass
    ctx["resumo_faturamento"] = (
        f"No período de {periodo_str} foram processadas {ctx['total_transacoes']} transações, "
        f"totalizando {ctx['faturamento_bruto']} em faturamento bruto. "
        f"O valor líquido recebido foi de {ctx['valor_liquido']} "
        f"({ctx['percentual_liquido']}% do bruto), com taxa média de {ctx['taxa_media']}%."
    )

    # Destaques
    destaques = []
    if total_transacoes:
        destaques.append({
            "tipo": "success",
            "icone": "📊",
            "titulo": "Volume processado",
            "descricao": f"{ctx['total_transacoes']} transações conciliadas no período.",
        })
    if total_divergencias_num > 0:
        percentual_div = (total_divergencias_num / faturamento_bruto * 100) if faturamento_bruto else 0.0
        destaques.append({
            "tipo": "warning",
            "icone": "⚠️",
            "titulo": "Divergências identificadas",
            "descricao": f"Foram identificadas perdas no valor de {ctx['total_divergencias']}, representando {_fmt_pct(percentual_div)}% do faturamento total.",
        })
    else:
        destaques.append({
            "tipo": "success",
            "icone": "✅",
            "titulo": "Sem divergências",
            "descricao": "Nenhuma perda estimada identificada no período.",
        })
    ctx["destaques"] = destaques

    # Conclusão
    if total_divergencias_num > 0:
        ctx["conclusao"] = (
            f"A análise identificou {ctx['total_divergencias']} em divergências financeiras. "
            "Recomenda-se revisão detalhada das transações marcadas e contestação junto à adquirente."
        )
        ctx["recomendacoes"] = [
            "Revisar as transações com divergência no Relatório Analítico.",
            "Acionar a adquirente para contestação dentro do prazo.",
            "Monitorar o próximo ciclo para verificar regularização.",
        ]
    else:
        ctx["conclusao"] = (
            "A conciliação foi concluída sem divergências no período analisado. "
            "Os valores cobrados estão de acordo com as taxas contratadas."
        )
        ctx["recomendacoes"] = []

    return ctx


def _gerar_excel_xlsx(
    dados: dict,
    processamento_id: str,
    engine: Engine,
    output_path: str,
    opcoes: Optional[dict] = None,
) -> None:
    """
    Gera arquivo .xlsx usando pd.ExcelWriter (muito mais rápido que openpyxl célula-a-célula).
    Abas: Vendas Completas, Perdas por Semestre, Taxas Min-Max,
          Contagem Transações, Dados Bancários,
          Top 3 Maiores Valores, Vendas Filtradas, Recebíveis Filtrados.
    """
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    opcoes = opcoes or {}
    adquirente_filtro = opcoes.get("adquirente") or None
    if adquirente_filtro and adquirente_filtro.lower() in ("todos", "all", ""):
        adquirente_filtro = None

    # Preparar DataFrames
    df_vc = dados.get("vendas_calculos", pd.DataFrame())
    if adquirente_filtro and not df_vc.empty:
        col_adq = next((c for c in df_vc.columns if "adquirente" in c.lower()), None)
        if col_adq:
            df_vc = df_vc[df_vc[col_adq].astype(str).str.lower() == adquirente_filtro.lower()]
    # Extrair stats reais ANTES de remover colunas auxiliares
    def _xreal(col, fallback=0.0):
        return df_vc[col].iloc[0] if not df_vc.empty and col in df_vc.columns else fallback
    _ex_n    = int(_xreal("_total_transacoes_real",   len(df_vc)))
    _ex_fat  = float(_xreal("_faturamento_bruto_real", 0.0))
    _ex_liq  = float(_xreal("_valor_liquido_real",     0.0))
    _ex_tm   = float(_xreal("_taxa_media_real",        0.0))
    _ex_tmin = float(_xreal("_taxa_min_real",          0.0))
    _ex_tmax = float(_xreal("_taxa_max_real",          0.0))

    # Remover colunas auxiliares internas antes de escrever no Excel
    _aux_cols = [c for c in df_vc.columns if c.startswith("_")]
    if _aux_cols:
        df_vc = df_vc.drop(columns=_aux_cols)

    # Top 3 maiores valores
    df_top3 = pd.DataFrame()
    if not df_vc.empty:
        col_bruto = next((c for c in df_vc.columns if c.lower() in ("vl_venda", "valor_venda", "vl_bruto")), None)
        if not col_bruto:
            col_bruto = next((c for c in df_vc.columns if c.lower().startswith("vl_") and "venda" in c.lower()), None)
        if col_bruto:
            df_top3 = df_vc.nlargest(3, col_bruto)

    # Agregar vendas_filtradas: Status | Bandeira | Forma de Pagamento | Qtd | Valor Bruto Total
    df_vf_raw = dados.get("vendas_filtradas", pd.DataFrame())
    if not df_vf_raw.empty:
        try:
            col_status_vf   = next((c for c in df_vf_raw.columns if "status" in c.lower()), None)
            col_bandeira_vf = next((c for c in df_vf_raw.columns if "bandeira" in c.lower()), None)
            col_forma_vf    = next((c for c in df_vf_raw.columns if "forma" in c.lower() and "pagamento" in c.lower()), None)
            col_val_vf      = next((c for c in df_vf_raw.columns if "valor" in c.lower() and "venda" in c.lower()), None)
            if not col_val_vf:
                col_val_vf = next((c for c in df_vf_raw.columns if c.lower().startswith("vl_") and "venda" in c.lower()), None)
            grp_cols_vf = [c for c in [col_status_vf, col_bandeira_vf, col_forma_vf] if c]
            if grp_cols_vf and col_val_vf:
                df_vf_excel = df_vf_raw.groupby(grp_cols_vf, dropna=False).agg(
                    **{"Quantidade de Vendas": (grp_cols_vf[0], "count"),
                       "Valor Bruto Total": (col_val_vf, "sum")}
                ).reset_index()
                rename_vf = {}
                if col_status_vf:   rename_vf[col_status_vf]   = "Status"
                if col_bandeira_vf: rename_vf[col_bandeira_vf] = "Bandeira"
                if col_forma_vf:    rename_vf[col_forma_vf]    = "Forma de Pagamento"
                df_vf_excel.rename(columns=rename_vf, inplace=True)
                total_vf = {c: "" for c in df_vf_excel.columns}
                total_vf["Status"] = "TOTAL GERAL"
                total_vf["Quantidade de Vendas"] = int(df_vf_excel["Quantidade de Vendas"].sum())
                total_vf["Valor Bruto Total"] = float(df_vf_excel["Valor Bruto Total"].sum())
                df_vf_excel = pd.concat([df_vf_excel, pd.DataFrame([total_vf])], ignore_index=True)
            else:
                df_vf_excel = df_vf_raw
        except Exception:
            df_vf_excel = df_vf_raw
    else:
        df_vf_excel = df_vf_raw

    # Agregar recebiveis_filtrados para evitar escrever 500K+ linhas no Excel
    df_rf_raw = dados.get("recebiveis_filtrados", pd.DataFrame())
    if not df_rf_raw.empty:
        try:
            grp_col_rf = next((c for c in df_rf_raw.columns if "lancamento" in c.lower()), None)
            col_vr_rf = next((c for c in df_rf_raw.columns if "valor_recebivel" in c.lower()), None)
            col_vl_rf = next((c for c in df_rf_raw.columns if "valor_liquido" in c.lower()), None)
            if grp_col_rf:
                agg_rf = {"Quantidade": (grp_col_rf, "count")}
                if col_vr_rf: agg_rf["Valor Recebível Total"] = (col_vr_rf, "sum")
                if col_vl_rf: agg_rf["Valor Líquido Total"] = (col_vl_rf, "sum")
                df_rf_excel = df_rf_raw.groupby(grp_col_rf).agg(**agg_rf).reset_index()
                df_rf_excel.rename(columns={grp_col_rf: "Tipo de Lançamento"}, inplace=True)
            else:
                df_rf_excel = df_rf_raw.head(5000)
        except Exception:
            df_rf_excel = df_rf_raw.head(5000)
    else:
        df_rf_excel = df_rf_raw

    # Aba de resumo com stats reais (extraídas antes do drop das colunas auxiliares)
    _pct_liq = (_ex_liq / _ex_fat * 100) if _ex_fat else 0.0
    df_resumo = pd.DataFrame([
        {"Campo": "Total de Transações",       "Valor": f"{_ex_n:,}".replace(",", ".")},
        {"Campo": "Faturamento Bruto",         "Valor": f"R$ {_ex_fat:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")},
        {"Campo": "Valor Líquido",             "Valor": f"R$ {_ex_liq:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")},
        {"Campo": "% Líquido / Bruto",         "Valor": f"{_pct_liq:.2f}%".replace(".", ",")},
        {"Campo": "Taxa Média (%)",            "Valor": f"{_ex_tm:.4f}%".replace(".", ",")},
        {"Campo": "Taxa Mínima (%)",           "Valor": f"{_ex_tmin:.4f}%".replace(".", ",")},
        {"Campo": "Taxa Máxima (%)",           "Valor": f"{_ex_tmax:.4f}%".replace(".", ",")},
        {"Campo": "Diferença de Taxa (%)",     "Valor": f"{_ex_tmax - _ex_tmin:.4f}%".replace(".", ",")},
        {"Campo": "Nota",                      "Valor": f"Amostra: 100.000 de {_ex_n:,} transações na aba seguinte".replace(",", ".")},
    ])

    abas = [
        ("0. Resumo",                     df_resumo),
        ("1. Amostra (100k vendas)",       df_vc),
        ("2. Sumário Descontos Semestre",  dados.get("perdas_semestre", pd.DataFrame())),
        ("3. Taxas Min-Max",               dados.get("taxas_minmax", pd.DataFrame())),
        ("4. Contagem Transações",         dados.get("contagem_transacoes", pd.DataFrame())),
        ("5. Dados Bancários",             dados.get("dados_bancarios", pd.DataFrame())),
        ("6. Top 3 Maiores Valores",       df_top3),
        ("7. Vendas Filtradas",            df_vf_excel),
        ("8. Recebíveis Filtrados",        df_rf_excel),
    ]

    HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
    HEADER_FONT = Font(bold=True, color="FFFFFF", size=10)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        abas_escritas = 0
        for nome_aba, df in abas:
            if df is None or df.empty:
                continue
            df.to_excel(writer, sheet_name=nome_aba, index=False)
            ws = writer.sheets[nome_aba]
            # Formatar cabeçalho
            for cell in ws[1]:
                cell.font = HEADER_FONT
                cell.fill = HEADER_FILL
                cell.alignment = Alignment(horizontal="center", wrap_text=True)
            # Ajustar largura das colunas
            for ci in range(1, len(df.columns) + 1):
                ws.column_dimensions[get_column_letter(ci)].width = 18
            abas_escritas += 1

        if abas_escritas == 0:
            # Garantir ao menos uma aba
            pd.DataFrame([{"info": "Nenhuma seção disponível"}]).to_excel(
                writer, sheet_name="Sem dados", index=False
            )


def _montar_contexto_html(dados: dict, processamento_id: str, engine: Engine, opcoes: Optional[dict] = None) -> dict:
    """
    Monta contexto completo para templates HTML analítico/sem_capa.
    Replica o que gerar_relatorio_html fazia: tabelas, gráficos, materialidade, imagens.
    """
    from modules.reports import (
        gerar_tabela_html,
        obter_dados_processamento,
        criar_tabela_sumario,
        criar_grafico_vendas_por_bandeira,
        criar_grafico_vendas_por_forma_pagamento,
        criar_grafico_vendas_por_mes,
        criar_grafico_valor_medio_por_bandeira,
        gerar_demonstrativo_vendas_filtradas,
        gerar_demonstrativo_recebiveis_filtrados,
    )

    opcoes = opcoes or {}
    incluir_filtradas = opcoes.get("incluir_filtradas", False)
    incluir_rec_filtrados = opcoes.get("incluir_recebiveis_filtrados", False)
    apenas_com_perdas = opcoes.get("apenas_com_perdas", False)
    adquirente_filtro = opcoes.get("adquirente") or None
    if adquirente_filtro and adquirente_filtro.lower() in ("todos", "all", ""):
        adquirente_filtro = None

    # Aplicar filtro de adquirente no df_vc (parquet já filtrado — mas garante consistência)
    df_vc = dados.get("vendas_calculos", pd.DataFrame())
    if adquirente_filtro and not df_vc.empty:
        col_adq = next((c for c in df_vc.columns if "adquirente" in c.lower()), None)
        if col_adq:
            df_vc = df_vc[df_vc[col_adq].astype(str).str.lower() == adquirente_filtro.lower()]

    if apenas_com_perdas and not df_vc.empty and "perda" in df_vc.columns:
        df_vc = df_vc[df_vc["perda"].fillna(0) != 0]

    ctx: dict = {
        "incluir_filtradas": incluir_filtradas,
        "incluir_recebiveis_filtrados": incluir_rec_filtrados,
        "apenas_com_perdas": apenas_com_perdas,
        "adquirente": adquirente_filtro or "Todos",
        "adquirente_principal": adquirente_filtro or "Todos",
        "data_geracao": datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        "disclaimer_text": "Todas as análises são baseadas exclusivamente nos dados fornecidos pela Adquirente.",
    }

    # ── Tabela Sumário (Informações do Processamento + Estatísticas Gerais) ──
    if not df_vc.empty:
        try:
            col_bruto = next((c for c in df_vc.columns if c.lower() in ("vl_venda", "valor_venda", "vl_bruto")), None)
            if not col_bruto:
                col_bruto = next((c for c in df_vc.columns if c.lower().startswith("vl_") and "venda" in c.lower()), None)
            col_taxa = next((c for c in df_vc.columns if c.lower() in ("tx_calc", "tx_venda")), None)
            col_data = next((c for c in df_vc.columns if "data" in c.lower()), None)

            vl = df_vc[col_bruto] if col_bruto else pd.Series(dtype=float)
            tx = df_vc[col_taxa] if col_taxa else pd.Series(dtype=float)
            valor_total = float(vl.sum()) if not vl.empty else 0.0
            datas = pd.to_datetime(df_vc[col_data], errors="coerce") if col_data else pd.Series(dtype="datetime64[ns]")
            primeira = datas.min() if not datas.empty else None
            ultima = datas.max() if not datas.empty else None
            periodo_dias = int((ultima - primeira).days) if primeira and ultima else 0

            def _rstat(col, fallback):
                return df_vc[col].iloc[0] if col in df_vc.columns else fallback
            _qtd_real   = int(_rstat("_total_transacoes_real",   len(df_vc)))
            _fat_real   = float(_rstat("_faturamento_bruto_real", vl.sum() if not vl.empty else 0.0))
            _vmedio     = float(_rstat("_valor_medio_real",       vl.mean() if not vl.empty else 0.0))
            _vmin       = float(_rstat("_valor_min_real",         vl.min() if not vl.empty else 0.0))
            _vmax       = float(_rstat("_valor_max_real",         vl.max() if not vl.empty else 0.0))
            _tmin       = float(_rstat("_taxa_min_real",          tx.min() if not tx.empty else 0.0))
            _tmax       = float(_rstat("_taxa_max_real",          tx.max() if not tx.empty else 0.0))
            valor_total = _fat_real
            estatisticas_sumario = {
                "quantidade": _qtd_real,
                "valor_total": _fat_real,
                "valor_medio": _vmedio,
                "valor_min": _vmin,
                "valor_max": _vmax,
                "diferenca_taxa": _tmax - _tmin,
                "primeira_venda": primeira.strftime("%d/%m/%Y") if primeira else "—",
                "ultima_venda": ultima.strftime("%d/%m/%Y") if ultima else "—",
                "periodo_dias": periodo_dias,
            }
            estatisticas_taxas = {
                "min_taxa": _tmin,
                "max_taxa": _tmax,
            }

            # Metadados do processamento via DB
            try:
                from modules.reports import obter_ecs_distintos_processamento
                _meta = obter_dados_processamento(engine, processamento_id)
            except Exception:
                _meta = {}

            metadados = {
                "id_processamento": processamento_id,
                "cliente_nome": _meta.get("cliente_nome", "—"),
                "ec_id": _meta.get("ec_id", "—"),
                "adquirente": adquirente_filtro or "Todos",
                "data_processamento": _meta.get("data_processamento", "—"),
            }
            try:
                ecs = obter_ecs_distintos_processamento(engine, processamento_id, adquirente_filtro)
            except Exception:
                ecs = []
            adqs = [adquirente_filtro] if adquirente_filtro else (_meta.get("adquirentes", ["—"]))

            ctx["tabela_sumario_html"] = criar_tabela_sumario(
                estatisticas_sumario, metadados, estatisticas_taxas, ecs, adqs
            )
            ctx.update(_meta)  # propaga cliente_nome, periodo, etc.
        except Exception as e:
            logger.warning(f"[HTML] tabela_sumario_html erro: {e}")
            ctx.setdefault("tabela_sumario_html", "")
    else:
        ctx.setdefault("tabela_sumario_html", "")

    # ── Tabelas padrão ──
    if "perdas_semestre" in dados and not dados["perdas_semestre"].empty:
        from modules.reports import sumarizar_perdas_por_semestre
        try:
            df_ps = sumarizar_perdas_por_semestre(dados["perdas_semestre"])
        except Exception:
            df_ps = dados["perdas_semestre"]
        ctx["tabela_perdas_semestre_html"] = gerar_tabela_html(
            df_ps, "Sumário de Registros com Descontos Contestáveis por Semestre"
        )
    else:
        ctx.setdefault("tabela_perdas_semestre_html", "")

    if "taxas_minmax" in dados and not dados["taxas_minmax"].empty:
        ctx["tabela_min_max_taxas_html"] = gerar_tabela_html(
            dados["taxas_minmax"], "Análise de Taxas Mínimas e Máximas por Semestre"
        )
    else:
        ctx.setdefault("tabela_min_max_taxas_html", "")

    if "contagem_transacoes" in dados and not dados["contagem_transacoes"].empty:
        ctx["tabela_contagem_taxas_html"] = gerar_tabela_html(
            dados["contagem_transacoes"], "Contagem de Transações por Ano-Semestre"
        )
    else:
        ctx.setdefault("tabela_contagem_taxas_html", "")

    if "recebiveis_sumario" in dados and not dados["recebiveis_sumario"].empty:
        try:
            df_rec_sum = dados["recebiveis_sumario"].copy()
            col_sem = next((c for c in df_rec_sum.columns if "semestre" in c.lower() or c == "Ano-Semestre"), "Ano-Semestre")
            col_val = next((c for c in df_rec_sum.columns if "valor" in c.lower()), None)
            col_lanc = next((c for c in df_rec_sum.columns if "lançamento" in c.lower() or "lancamento" in c.lower()), None)
            if col_val and col_sem:
                rows_with_sub = []
                for sem, grp in df_rec_sum.groupby(col_sem, sort=True):
                    for _, r in grp.iterrows():
                        rows_with_sub.append(r.to_dict())
                    sub_val = grp[col_val].sum()
                    sub_row = {c: "" for c in df_rec_sum.columns}
                    sub_row[col_sem] = f"SUBTOTAL {sem}"
                    sub_row[col_val] = sub_val
                    rows_with_sub.append(sub_row)
                total_row = {c: "" for c in df_rec_sum.columns}
                total_row[col_sem] = "TOTAL GERAL"
                total_row[col_val] = df_rec_sum[col_val].sum()
                rows_with_sub.append(total_row)
                df_rec_sum = pd.DataFrame(rows_with_sub, columns=df_rec_sum.columns)
            ctx["tabela_sumario_recebiveis_html"] = gerar_tabela_html(
                df_rec_sum, "Sumário de Recebíveis com Descontos Contestáveis"
            )
        except Exception as e:
            logger.warning(f"[HTML] tabela_sumario_recebiveis_html erro: {e}")
            ctx["tabela_sumario_recebiveis_html"] = gerar_tabela_html(
                dados["recebiveis_sumario"], "Sumário de Recebíveis com Descontos Contestáveis"
            )
    else:
        ctx.setdefault("tabela_sumario_recebiveis_html", "")

    if "dados_bancarios" in dados and not dados["dados_bancarios"].empty:
        ctx["tabela_dados_bancarios_html"] = gerar_tabela_html(
            dados["dados_bancarios"], "Dados Bancários Distintos"
        )
    else:
        ctx.setdefault("tabela_dados_bancarios_html", "")

    # ── Gráficos Plotly ──
    # As funções de gráfico esperam nomes de colunas com capitalização específica
    _col_map = {
        "bandeira": "Bandeira",
        "forma_pagamento": "Forma_de_pagamento",
        "data_venda": "Data_da_venda",
        "vl_venda": "Valor_da_venda",
    }
    df_vc_charts = df_vc.rename(columns={k: v for k, v in _col_map.items() if k in df_vc.columns}) if not df_vc.empty else df_vc
    for fn, key in [
        (criar_grafico_vendas_por_bandeira,       "grafico_bandeiras_html"),
        (criar_grafico_vendas_por_forma_pagamento, "grafico_forma_pagamento_html"),
        (criar_grafico_vendas_por_mes,             "grafico_meses_html"),
        (criar_grafico_valor_medio_por_bandeira,   "grafico_valores_html"),
    ]:
        try:
            ctx[key] = fn(df_vc_charts) if not df_vc_charts.empty else ""
        except Exception as e:
            logger.warning(f"[HTML] {key} erro: {e}")
            ctx[key] = ""

    # ── Materialidade ──
    # Total = Perda Monetária MDR + Perda Monetária RR/RA (de perdas_semestre)
    # recebiveis_sumario tem valor BRUTO — não entra na materialidade
    try:
        def _parse_brl(v) -> float:
            try:
                return abs(float(str(v).replace("R$", "").replace(".", "").replace(",", ".").strip()))
            except Exception:
                return 0.0

        perda_total = 0.0
        df_ps_mat = dados.get("perdas_semestre", pd.DataFrame())
        if df_ps_mat is not None and not df_ps_mat.empty:
            col_sem_mat = next((c for c in df_ps_mat.columns if "semestre" in c.lower() or c == "Ano-Semestre"), None)
            df_mat_rows = df_ps_mat[~df_ps_mat[col_sem_mat].astype(str).str.contains("TOTAL", case=False, na=False)] if col_sem_mat else df_ps_mat
            col_mdr = next((c for c in df_mat_rows.columns if "mdr" in c.lower()), None)
            if col_mdr:
                perda_total += sum(_parse_brl(v) for v in df_mat_rows[col_mdr])
            col_rr = next((c for c in df_mat_rows.columns if ("rr" in c.lower() or "ra" in c.lower()) and "perda" in c.lower()), None)
            if col_rr:
                perda_total += sum(_parse_brl(v) for v in df_mat_rows[col_rr])

        # Adicionar total de recebíveis com descontos contestáveis
        df_rec_mat = dados.get("recebiveis_sumario", pd.DataFrame())
        if df_rec_mat is not None and not df_rec_mat.empty:
            col_sem_rec = next((c for c in df_rec_mat.columns if "semestre" in c.lower() or c == "Ano-Semestre"), None)
            df_rec_rows = df_rec_mat[~df_rec_mat[col_sem_rec].astype(str).str.contains("TOTAL", case=False, na=False)] if col_sem_rec else df_rec_mat
            col_val_rec = next((c for c in df_rec_rows.columns if "valor" in c.lower()), None)
            if col_val_rec:
                rec_vals = pd.to_numeric(df_rec_rows[col_val_rec], errors="coerce").fillna(0)
                perda_total += float(rec_vals.sum())

        # Fallback se parquets não disponíveis
        if perda_total == 0.0 and not df_vc.empty:
            mdr = abs(float(df_vc["perda"].fillna(0).sum())) if "perda" in df_vc.columns else 0.0
            rr  = abs(float(df_vc["perda_rr"].fillna(0).sum())) if "perda_rr" in df_vc.columns else 0.0
            perda_total = mdr + rr

        valor_total_ctx = float(df_vc[col_bruto].sum()) if (not df_vc.empty and col_bruto) else 1.0
        pct = (perda_total / valor_total_ctx * 100) if valor_total_ctx else 0.0
        ctx["materialidade_valor"] = f"R$ {perda_total:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        ctx["materialidade_percentual"] = f"{pct:.2f}%".replace(".", ",")
    except Exception as e:
        logger.warning(f"[HTML] materialidade erro: {e}")
        ctx.setdefault("materialidade_valor", "R$ 0,00")
        ctx.setdefault("materialidade_percentual", "0,00%")

    # ── Tabelas condicionais ──
    if incluir_filtradas:
        df_vf = dados.get("vendas_filtradas", pd.DataFrame())
        if not df_vf.empty:
            try:
                col_status   = next((c for c in df_vf.columns if "status" in c.lower()), None)
                col_bandeira = next((c for c in df_vf.columns if "bandeira" in c.lower()), None)
                col_forma    = next((c for c in df_vf.columns if "forma" in c.lower() and "pagamento" in c.lower()), None)
                col_val      = next((c for c in df_vf.columns if "valor" in c.lower() and "venda" in c.lower()), None)
                if not col_val:
                    col_val = next((c for c in df_vf.columns if c.lower().startswith("vl_") and "venda" in c.lower()), None)
                grp_cols = [c for c in [col_status, col_bandeira, col_forma] if c]
                if grp_cols and col_val:
                    df_agg_vf = df_vf.groupby(grp_cols, dropna=False).agg(
                        **{"Quantidade de Vendas": (grp_cols[0], "count"),
                           "Valor Bruto Total": (col_val, "sum")}
                    ).reset_index()
                    rename_map = {}
                    if col_status:   rename_map[col_status]   = "Status"
                    if col_bandeira: rename_map[col_bandeira] = "Bandeira"
                    if col_forma:    rename_map[col_forma]    = "Forma de Pagamento"
                    df_agg_vf.rename(columns=rename_map, inplace=True)
                    # Linha TOTAL
                    total_row = {c: "" for c in df_agg_vf.columns}
                    total_row["Status"] = "TOTAL GERAL"
                    total_row["Quantidade de Vendas"] = int(df_agg_vf["Quantidade de Vendas"].sum())
                    total_row["Valor Bruto Total"] = float(df_agg_vf["Valor Bruto Total"].sum())
                    df_agg_vf = pd.concat([df_agg_vf, pd.DataFrame([total_row])], ignore_index=True)
                    ctx["tabela_vendas_filtradas_html"] = gerar_tabela_html(df_agg_vf, "Demonstrativo de Outras Vendas")
                else:
                    ctx["tabela_vendas_filtradas_html"] = gerar_tabela_html(df_vf.head(200), "Demonstrativo de Outras Vendas")
            except Exception as e:
                logger.warning(f"[HTML] tabela_vendas_filtradas_html erro: {e}")
                ctx["tabela_vendas_filtradas_html"] = ""
        else:
            ctx["tabela_vendas_filtradas_html"] = ""

    if incluir_rec_filtrados:
        df_rf = dados.get("recebiveis_filtrados", pd.DataFrame())
        if not df_rf.empty:
            try:
                # Agregar por lancamento — evita renderizar 500K+ linhas como HTML
                grp_col = next((c for c in df_rf.columns if "lancamento" in c.lower()), None)
                col_vr = next((c for c in df_rf.columns if "valor_recebivel" in c.lower()), None)
                col_vl = next((c for c in df_rf.columns if "valor_liquido" in c.lower()), None)
                if grp_col:
                    agg_spec = {"Quantidade": (grp_col, "count")}
                    if col_vr: agg_spec["Valor Recebível Total"] = (col_vr, "sum")
                    if col_vl: agg_spec["Valor Líquido Total"] = (col_vl, "sum")
                    df_agg_rf = df_rf.groupby(grp_col).agg(**agg_spec).reset_index()
                    df_agg_rf.rename(columns={grp_col: "Tipo de Lançamento"}, inplace=True)
                    ctx["tabela_recebiveis_filtrados_html"] = gerar_tabela_html(df_agg_rf, "Demonstrativo de Outros Registros de Desconto")
                else:
                    ctx["tabela_recebiveis_filtrados_html"] = gerar_tabela_html(df_rf.head(200), "Demonstrativo de Outros Registros de Desconto")
            except Exception as e:
                logger.warning(f"[HTML] tabela_recebiveis_filtrados_html erro: {e}")
                ctx["tabela_recebiveis_filtrados_html"] = ""
        else:
            ctx["tabela_recebiveis_filtrados_html"] = ""

    # ── Evidências ──
    evidencias = dados.get("evidencias", {}) or {}
    for ev_key, titulo in [
        ("maiores_valores",  "Evidências — Maiores Valores"),
        ("menores_valores",  "Evidências — Menores Valores"),
        ("maiores_taxas",    "Evidências — Maiores Taxas"),
        ("menores_taxas",    "Evidências — Menores Taxas"),
    ]:
        ctx_key = f"tabela_evidencias_{ev_key}_html"
        try:
            df_ev = evidencias.get(ev_key)
            if df_ev is not None and not df_ev.empty:
                ctx[ctx_key] = gerar_tabela_html(df_ev, titulo)
            else:
                ctx[ctx_key] = ""
        except Exception:
            ctx[ctx_key] = ""

    # ── Imagens (capa e cabeçalho) ──
    _assets = os.path.normpath(os.path.join(
        os.path.dirname(__file__), "..", "..", "..", "..", "assets"
    ))
    for img_key, filename in [("cover_image_path", "capa_relatorio.jpg"), ("header_image_path", "cabecalho_financial.png")]:
        img_path = os.path.join(_assets, filename)
        if os.path.exists(img_path):
            try:
                import base64 as _b64
                with open(img_path, "rb") as _f:
                    _ext = "jpeg" if filename.endswith(".jpg") else "png"
                    ctx[img_key] = f"data:image/{_ext};base64,{_b64.b64encode(_f.read()).decode()}"
            except Exception:
                ctx[img_key] = ""
        else:
            ctx[img_key] = ""

    return ctx
