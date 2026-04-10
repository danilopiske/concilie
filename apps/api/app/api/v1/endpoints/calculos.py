import io
import logging
from collections import defaultdict
from decimal import Decimal
from typing import List

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from sqlalchemy.orm import Session

logger = logging.getLogger(__name__)
from fastapi import BackgroundTasks

from app.core.database import get_db
from app.repositories.calculo_repository import CalculoRepository
from app.schemas.calculo import (
    AnalisePeriodosResponse,
    CalculoHistoryItem,
    CalculoPreviewRequest,
    CalculoRequest,
    CalculoResultado,
    CalculoStats,
)
from app.services.calculo_service import CalculoService
from app.services.reconciliation_core import ReconciliationCore

router = APIRouter()

@router.get("/analise-periodos/{processamento_id:path}", response_model=AnalisePeriodosResponse)
def analise_periodos(
    processamento_id: str,
    threshold: float = Query(0.5, ge=0.0, le=1.0),
    db: Session = Depends(get_db),
):
    print(f"DEBUG: analise_periodos for {processamento_id} (threshold={threshold})")
    repo = CalculoRepository(db)
    return repo.analisar_periodos(processamento_id, threshold)



@router.post("/preview", response_model=CalculoStats)
def preview_calculo(req: CalculoPreviewRequest, db: Session = Depends(get_db)):
    repo = CalculoRepository(db)
    return repo.preview_calculo(req)

@router.get("/historico-calculos", response_model=List[CalculoHistoryItem])
async def get_historico_calculos(
    skip: int = Query(0),
    limit: int = Query(50),
    db: Session = Depends(get_db)
):
    """
    Retorna o histórico de cálculos únicos salvos no banco.
    """
    repo = CalculoRepository(db)
    return repo.listar_historico(skip=skip, limit=limit)

@router.post("/processar")
def processar_calculo(req: CalculoRequest, db: Session = Depends(get_db)):
    """
    Runs calculation synchronously using the high-performance Polars engine.
    For very large datasets, use /processar-async.
    """
    try:
        repo = CalculoRepository(db)
        custom_id = repo.processar_calculo(req)

        engine = db.get_bind()
        result = ReconciliationCore.calculate_rates(
            engine=engine,
            proc_id=req.processamento_id,
            tipo_taxa=req.tipo_taxa,
            usar_taxa_cad=req.usar_taxa_cad,
            tem_receba_rapido=req.tem_receba_rapido,
            custom_calc_id=custom_id
        )
        if not result.get("success"):
            raise HTTPException(status_code=500, detail=result.get("error"))

        return {
            "status": "success",
            "message": f"Cálculo processado com sucesso. {result.get('rows')} registros.",
            "calc_id": custom_id
        }
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@router.post("/processar-async")
async def processar_calculo_async(
    req: CalculoRequest,
    background_tasks: BackgroundTasks,
    usuario: str = "api_user",
    db: Session = Depends(get_db)
):
    """Starts calculation in background and returns a task_id."""
    service = CalculoService(db)
    task = service.create_calculo_task(
        processamento_id=req.processamento_id,
        tipo_taxa=req.tipo_taxa,
        usuario=usuario,
        usar_taxa_cad=req.usar_taxa_cad,
        tem_receba_rapido=req.tem_receba_rapido,
        substituir=req.substituir
    )

    background_tasks.add_task(service.run_async_calculo, task.id)

    return {
        "status": "processing",
        "task_id": task.id,
        "message": "Cálculo iniciado em segundo plano."
    }

@router.get("/task/{task_id:path}")
async def get_task_status(
    task_id: str,
    db: Session = Depends(get_db)
):
    """Returns current calculation task status."""
    service = CalculoService(db)
    task = service.get_task(task_id)
    if not task:
        raise HTTPException(status_code=404, detail="Tarefa não encontrada")

    return {
        "id": task.id,
        "status": task.status,
        "progress": task.progress,
        "message": task.message,
        "updated_at": task.updated_at,
        "processamento_id": task.processamento_id,
        "tipo_taxa": task.tipo_taxa
    }

@router.get("/resultados/{calc_id:path}", response_model=List[CalculoResultado])
def listar_resultados(
    calc_id: str,
    skip: int = Query(0),
    limit: int = Query(100),
    db: Session = Depends(get_db)
):
    logger.debug("Buscando resultados para calc_id: %s", calc_id)
    repo = CalculoRepository(db)
    return repo.listar_resultados(calc_id, skip, limit)


@router.get("/export/{calc_id:path}")
def export_calculo_excel(calc_id: str, db: Session = Depends(get_db)):
    """Exporta todos os resultados de um cálculo para Excel (6 sheets analíticas)."""
    from app.models.vendas_calculos import VendasCalculos

    registros = (
        db.query(VendasCalculos)
        .filter(VendasCalculos.calc_id == calc_id)
        .order_by(VendasCalculos.perda.asc())
        .all()
    )
    if not registros:
        raise HTTPException(status_code=404, detail="Nenhum resultado encontrado para este cálculo.")

    wb = Workbook()

    # ── Cores e estilos ─────────────────────────────────────────────────────
    HEADER_FILL = PatternFill("solid", fgColor="223A6B")
    HEADER_FONT = Font(bold=True, color="FFFFFF")
    LOSS_FILL   = PatternFill("solid", fgColor="FFE5E5")
    TOTAL_FONT  = Font(bold=True)

    def _style_header(ws, row=1):
        for cell in ws[row]:
            cell.fill = HEADER_FILL
            cell.font = HEADER_FONT
            cell.alignment = Alignment(horizontal="center")

    def _autofit(ws):
        for col_cells in ws.columns:
            length = max((len(str(c.value or "")) for c in col_cells), default=10)
            ws.column_dimensions[get_column_letter(col_cells[0].column)].width = min(length + 4, 40)

    # ── Sheet 1: Resumo ──────────────────────────────────────────────────────
    ws_resumo = wb.active
    ws_resumo.title = "Resumo"

    total_valor   = sum(float(r.vl_venda or 0) for r in registros)
    total_perda   = sum(float(r.perda or 0) for r in registros)
    com_perda     = sum(1 for r in registros if r.perda is not None and r.perda < 0)
    media_tx_vend = sum(float(r.tx_venda or 0) for r in registros) / len(registros)
    media_tx_calc = (
        sum(float(r.tx_calc or 0) for r in registros if r.tx_calc is not None)
        / max(sum(1 for r in registros if r.tx_calc is not None), 1)
    )

    resumo_rows = [
        ("Parâmetro", "Valor"),
        ("Cálculo ID", registros[0].calc_id),
        ("Tipo de Taxa", registros[0].calc_tipo or "-"),
        ("Usuário", registros[0].calc_usuario or "-"),
        ("Data do Cálculo", str(registros[0].calc_data)[:19] if registros[0].calc_data else "-"),
        ("", ""),
        ("Total de Registros", len(registros)),
        ("Total Valor de Venda (R$)", round(total_valor, 2)),
        ("Total Perda/Ganho (R$)", round(total_perda, 2)),
        ("Registros com Perda", com_perda),
        ("Média Taxa Cobrada (%)", round(media_tx_vend, 4)),
        ("Média Taxa Calculada (%)", round(media_tx_calc, 4)),
    ]

    for row in resumo_rows:
        ws_resumo.append(list(row))

    _style_header(ws_resumo, row=1)
    _autofit(ws_resumo)

    # ── Sheet 2: Detalhamento ────────────────────────────────────────────────
    ws_det = wb.create_sheet("Detalhamento")

    headers = [
        "Data Venda", "Bandeira", "Forma Pagamento", "Adquirente",
        "NSU", "Cód. Autorização", "Vl. Venda (R$)",
        "Tx. Cobrada (%)", "Desc. Cobrado (R$)", "Vl. Líq. Cobrado (R$)",
        "Tx. Calculada (%)", "Desc. Calculado (R$)", "Vl. Líq. Calculado (R$)",
        "Diferença Taxa (%)", "Perda/Ganho (R$)",
    ]
    ws_det.append(headers)
    _style_header(ws_det, row=1)

    for r in registros:
        diff_taxa = (
            float(r.tx_venda) - float(r.tx_calc)
            if r.tx_venda is not None and r.tx_calc is not None
            else None
        )
        row_data = [
            str(r.data_venda)[:10] if r.data_venda else "",
            r.bandeira or "",
            r.forma_pagamento or "",
            r.adquirente or "",
            r.nsu or "",
            r.cod_autorizacao or "",
            float(r.vl_venda) if r.vl_venda is not None else None,
            float(r.tx_venda) if r.tx_venda is not None else None,
            float(r.desc_venda) if r.desc_venda is not None else None,
            float(r.vl_liq_venda) if r.vl_liq_venda is not None else None,
            float(r.tx_calc) if r.tx_calc is not None else None,
            float(r.desc_calc) if r.desc_calc is not None else None,
            float(r.vl_liq_calc) if r.vl_liq_calc is not None else None,
            round(diff_taxa, 4) if diff_taxa is not None else None,
            float(r.perda) if r.perda is not None else None,
        ]
        ws_det.append(row_data)
        # Destacar linhas com perda
        if r.perda is not None and float(r.perda) < 0:
            for cell in ws_det[ws_det.max_row]:
                cell.fill = LOSS_FILL

    # Linha de totais
    total_row = ws_det.max_row + 1
    ws_det.cell(total_row, 1, "TOTAL")
    ws_det.cell(total_row, 7, round(total_valor, 2))
    ws_det.cell(total_row, 15, round(total_perda, 2))
    for col in [1, 7, 15]:
        ws_det.cell(total_row, col).font = TOTAL_FONT

    _autofit(ws_det)

    # ── Sheet 3: Por Bandeira ────────────────────────────────────────────────
    def _build_group_sheet(ws, group_key_fn, col_headers):
        groups: dict = defaultdict(lambda: {"count": 0, "vl_venda": 0.0, "tx_venda_sum": 0.0, "tx_calc_sum": 0.0, "tx_calc_n": 0, "perda": 0.0, "com_perda": 0})
        for r in registros:
            key = group_key_fn(r)
            g = groups[key]
            g["count"] += 1
            g["vl_venda"] += float(r.vl_venda or 0)
            g["tx_venda_sum"] += float(r.tx_venda or 0)
            if r.tx_calc is not None:
                g["tx_calc_sum"] += float(r.tx_calc)
                g["tx_calc_n"] += 1
            g["perda"] += float(r.perda or 0)
            if r.perda is not None and float(r.perda) < 0:
                g["com_perda"] += 1

        ws.append(col_headers + ["Qtd Vendas", "Vl. Total (R$)", "Tx. Média Cobrada (%)", "Tx. Média Calculada (%)", "Perda Total (R$)", "Qtd c/ Perda"])
        _style_header(ws, row=1)

        for key in sorted(groups.keys(), key=lambda k: groups[k]["perda"]):
            g = groups[key]
            avg_tx_venda = round(g["tx_venda_sum"] / g["count"], 4) if g["count"] else 0
            avg_tx_calc = round(g["tx_calc_sum"] / g["tx_calc_n"], 4) if g["tx_calc_n"] else None
            row_vals = (list(key) if isinstance(key, tuple) else [key]) + [
                g["count"],
                round(g["vl_venda"], 2),
                avg_tx_venda,
                avg_tx_calc,
                round(g["perda"], 2),
                g["com_perda"],
            ]
            ws.append(row_vals)
            if g["perda"] < 0:
                for cell in ws[ws.max_row]:
                    cell.fill = LOSS_FILL

        # Linha de totais
        tr = ws.max_row + 1
        n_key_cols = 2 if isinstance(next(iter(groups.keys()), None), tuple) else 1
        ws.cell(tr, 1, "TOTAL").font = TOTAL_FONT
        ws.cell(tr, n_key_cols + 1, sum(g["count"] for g in groups.values())).font = TOTAL_FONT
        ws.cell(tr, n_key_cols + 2, round(sum(g["vl_venda"] for g in groups.values()), 2)).font = TOTAL_FONT
        ws.cell(tr, n_key_cols + 5, round(sum(g["perda"] for g in groups.values()), 2)).font = TOTAL_FONT

        _autofit(ws)

    ws_band = wb.create_sheet("Por Bandeira")
    _build_group_sheet(ws_band, lambda r: r.bandeira or "Sem Bandeira", ["Bandeira"])

    # ── Sheet 4: Por Forma de Pagamento ──────────────────────────────────────
    ws_fp = wb.create_sheet("Por Forma de Pgto")
    _build_group_sheet(ws_fp, lambda r: r.forma_pagamento or "Sem Forma Pgto", ["Forma de Pagamento"])

    # ── Sheet 5: Por Bandeira × Forma de Pagamento ───────────────────────────
    ws_cross = wb.create_sheet("Bandeira × Forma Pgto")
    _build_group_sheet(
        ws_cross,
        lambda r: (r.bandeira or "Sem Bandeira", r.forma_pagamento or "Sem Forma Pgto"),
        ["Bandeira", "Forma de Pagamento"],
    )

    # ── Sheet 6: Por Período (Mês) ────────────────────────────────────────────
    ws_mes = wb.create_sheet("Por Período (Mês)")
    _build_group_sheet(
        ws_mes,
        lambda r: str(r.data_venda)[:7] if r.data_venda else "Sem Data",
        ["Mês"],
    )

    # ── Helper: semestre (yyyy-S) a partir da data ───────────────────────────
    def _semestre(r):
        if not r.data_venda:
            return "Sem Data"
        d = str(r.data_venda)[:10]
        year, month = d[:4], int(d[5:7])
        return f"{year}-{'1' if month <= 6 else '2'}"

    # ── Sheet 7: Perdas por Semestre (estilo PDF) ─────────────────────────────
    ws_sem = wb.create_sheet("Perdas por Semestre")
    sem_groups: dict = defaultdict(lambda: {"vl_venda": 0.0, "perda_mdr": 0.0, "perda": 0.0})
    for r in registros:
        key = _semestre(r)
        g = sem_groups[key]
        g["vl_venda"] += float(r.vl_venda or 0)
        g["perda"] += float(r.perda or 0)
        # perda_rr não existe em VendasCalculos — perda já é total MDR
        g["perda_mdr"] += float(r.perda or 0)

    ws_sem.append(["Ano-Semestre", "Faturamento Bruto (R$)", "Perda Monetária MDR (R$)", "Perda RR/RA (R$)", "Perda Total (R$)", "% Perda"])
    _style_header(ws_sem, row=1)
    for key in sorted(sem_groups.keys()):
        g = sem_groups[key]
        pct = round(g["perda"] / g["vl_venda"] * 100, 2) if g["vl_venda"] else 0
        row_data = [key, round(g["vl_venda"], 2), round(g["perda_mdr"], 2), 0.0, round(g["perda"], 2), pct]
        ws_sem.append(row_data)
        if g["perda"] < 0:
            for cell in ws_sem[ws_sem.max_row]:
                cell.fill = LOSS_FILL
    tr = ws_sem.max_row + 1
    ws_sem.cell(tr, 1, "TOTAL").font = TOTAL_FONT
    ws_sem.cell(tr, 2, round(sum(g["vl_venda"] for g in sem_groups.values()), 2)).font = TOTAL_FONT
    ws_sem.cell(tr, 5, round(sum(g["perda"] for g in sem_groups.values()), 2)).font = TOTAL_FONT
    _autofit(ws_sem)

    # ── Sheet 8: Taxas Min/Max por Semestre ───────────────────────────────────
    ws_tmm = wb.create_sheet("Taxas Min-Max por Semestre")
    tmm_groups: dict = defaultdict(lambda: {"min_tx": None, "max_tx": None})
    for r in registros:
        key = (_semestre(r), r.bandeira or "Sem Bandeira", r.forma_pagamento or "Sem Forma")
        g = tmm_groups[key]
        tx = float(r.tx_venda) if r.tx_venda is not None else None
        if tx is not None:
            g["min_tx"] = tx if g["min_tx"] is None else min(g["min_tx"], tx)
            g["max_tx"] = tx if g["max_tx"] is None else max(g["max_tx"], tx)

    ws_tmm.append(["Ano-Semestre", "Bandeira", "Forma de Pagamento", "Taxa Mín (%)", "Taxa Máx (%)"])
    _style_header(ws_tmm, row=1)
    for key in sorted(tmm_groups.keys()):
        g = tmm_groups[key]
        ws_tmm.append([key[0], key[1], key[2],
                        round(g["min_tx"], 2) if g["min_tx"] is not None else None,
                        round(g["max_tx"], 2) if g["max_tx"] is not None else None])
    _autofit(ws_tmm)

    # ── Sheet 9: Contagem por Semestre × Bandeira × Modalidade ───────────────
    ws_cnt = wb.create_sheet("Contagem por Semestre")
    cnt_groups: dict = defaultdict(int)
    for r in registros:
        key = (_semestre(r), r.bandeira or "Sem Bandeira", r.forma_pagamento or "Sem Forma")
        cnt_groups[key] += 1

    ws_cnt.append(["Ano-Semestre", "Bandeira", "Forma de Pagamento", "Contagem"])
    _style_header(ws_cnt, row=1)
    for key in sorted(cnt_groups.keys()):
        ws_cnt.append([key[0], key[1], key[2], cnt_groups[key]])
    tr = ws_cnt.max_row + 1
    ws_cnt.cell(tr, 1, "TOTAL").font = TOTAL_FONT
    ws_cnt.cell(tr, 4, sum(cnt_groups.values())).font = TOTAL_FONT
    _autofit(ws_cnt)

    # ── Sheets 10-11: Recebíveis (via processamento_id) ──────────────────────
    proc_id = str(registros[0].processamento_id) if registros else None
    if proc_id:
        from app.models.recebiveis import Recebivel

        recebiveis = (
            db.query(Recebivel)
            .filter(Recebivel.processamentoid == proc_id)
            .all()
        )

        # Sheet 10: Sumário de Descontos Contestáveis por Semestre
        if recebiveis:
            ws_rec = wb.create_sheet("Recebíveis por Semestre")
            rec_groups: dict = defaultdict(lambda: defaultdict(float))
            for rv in recebiveis:
                if rv.data_recebivel:
                    m = rv.data_recebivel.month
                    sem_key = f"{rv.data_recebivel.year}-{'1' if m <= 6 else '2'}"
                else:
                    sem_key = "Sem Data"
                lancamento = rv.lancamento or "Outros"
                rec_groups[sem_key][lancamento] += float(rv.valor_recebivel or 0)

            ws_rec.append(["Ano-Semestre", "Lançamento", "Valor Total (R$)"])
            _style_header(ws_rec, row=1)
            SUBTOTAL_FILL = PatternFill("solid", fgColor="DDEEFF")
            for sem in sorted(rec_groups.keys()):
                subtotal = 0.0
                for lanc, valor in sorted(rec_groups[sem].items()):
                    ws_rec.append([sem, lanc, round(valor, 2)])
                    if valor < 0:
                        for cell in ws_rec[ws_rec.max_row]:
                            cell.fill = LOSS_FILL
                    subtotal += valor
                sub_row = ws_rec.max_row + 1
                ws_rec.cell(sub_row, 1, f"Subtotal {sem}").font = TOTAL_FONT
                ws_rec.cell(sub_row, 3, round(subtotal, 2)).font = TOTAL_FONT
                for cell in ws_rec[sub_row]:
                    cell.fill = SUBTOTAL_FILL
            total_rec = sum(
                v for sem in rec_groups.values() for v in sem.values()
            )
            tr = ws_rec.max_row + 1
            ws_rec.cell(tr, 1, "TOTAL GERAL").font = TOTAL_FONT
            ws_rec.cell(tr, 3, round(total_rec, 2)).font = TOTAL_FONT
            _autofit(ws_rec)

        # Sheet 11: Dados Bancários Distintos nos Recebíveis
        if recebiveis:
            bancos_vistos: set = set()
            banco_rows = []
            for rv in recebiveis:
                key = (rv.banco or "", rv.agencia or "", rv.conta or "")
                if key not in bancos_vistos and any(key):
                    bancos_vistos.add(key)
                    banco_rows.append(key)

            if banco_rows:
                ws_banco = wb.create_sheet("Dados Bancários")
                ws_banco.append(["Banco", "Agência", "Conta-Corrente"])
                _style_header(ws_banco, row=1)
                for row in sorted(banco_rows):
                    ws_banco.append(list(row))
                _autofit(ws_banco)

    # ── Serializar e retornar ────────────────────────────────────────────────
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"calculo_{calc_id.replace('/', '_')}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.delete("/deletar/{calc_id:path}")
async def delete_calculo(
    calc_id: str,
    db: Session = Depends(get_db)
):
    """
    Deleta um cálculo específico do banco de dados.
    """
    repo = CalculoRepository(db)
    repo.deletar_calculo(calc_id)
    return {"status": "success", "message": f"Cálculo {calc_id} deletado."}

@router.get("/historico")
async def get_historico(
    processamento_id: str = None,
    skip: int = 0,
    limit: int = 50,
    db: Session = Depends(get_db)
):
    """
    Retorna o histórico de TAREFAS de cálculo (async).
    """
    service = CalculoService(db)
    tasks = service.list_tasks(skip=skip, limit=limit, processamento_id=processamento_id)
    return tasks
